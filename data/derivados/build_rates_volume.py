#!/usr/bin/env python3
"""Construye la capa auditada de tasas y el efecto antes/después del dashboard v122.

La carpeta ``data`` del repositorio es la equivalencia local de ``/data`` en este
entorno Windows. El script nunca busca insumos fuera de esa carpeta.

Dependencia para leer el XLSX PNFC: openpyxl.
"""

from __future__ import annotations

import csv
import hashlib
import json
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


DATA_DIR = Path(__file__).resolve().parents[1]
SOURCE_DIR = DATA_DIR / "fuentes"
DERIVED_DIR = DATA_DIR / "derivados"
INPUT_HTML = DATA_DIR / "dashboard_kawaii_119_tasas_en_pesos.html"
OUTPUT_HTML = DATA_DIR / "dashboard_kawaii_122_efecto_tasas_antes_despues.html"
ROOT_OUTPUT_HTML = DATA_DIR.parent / "dashboard_kawaii_122_efecto_tasas_antes_despues.html"

BCRA_DIR = SOURCE_DIR / "tasas" / "bcra"
INDEC_DIR = SOURCE_DIR / "tasas" / "indec"
PNFC_DIR = SOURCE_DIR / "tasas" / "pnfc"
METHOD_DIR = SOURCE_DIR / "tasas" / "metodologia"

PERSONAL_CODES = {1936, 1938}
PF_CODES = {1307, 1309, 1311, 1313, 1315, 1317, 1319, 1321}
POST_START = "2023-12"
POST_END = "2026-07"
MIRROR_START = "2021-04"
MIRROR_END = "2023-11"


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def parse_number(value: str | int | float | None) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = value.strip()
    if not text or text.upper() == "NA":
        return None
    return float(text.replace(".", "").replace(",", ".")) if "," in text else float(text)


def js_json_array(html: str, name: str) -> list[dict[str, Any]]:
    match = re.search(rf"const\s+{re.escape(name)}\s*=\s*(\[.*?\]);", html, re.S)
    if not match:
        raise RuntimeError(f"No se encontró const {name} en el v119")
    return json.loads(match.group(1))


def annual_real_approx(row: dict[str, Any], kind: str) -> float:
    inflation_monthly = ((1 + row["inflacion"] / 100) ** (1 / 12) - 1) * 100
    nominal_monthly = (
        row["banco"] * 30.41666 / 365 if kind == "bank" else row["pf"] * 30 / 365
    )
    return ((1 + nominal_monthly / 100) / (1 + inflation_monthly / 100) - 1) * 100


def historical_baselines(
    annual: list[dict[str, Any]], modern: list[dict[str, Any]]
) -> dict[str, float | int]:
    bank_sum = 0.0
    pf_sum = 0.0
    months = 0
    for row in annual:
        if row["year"] <= 2018:
            bank_sum += annual_real_approx(row, "bank") * 12
            pf_sum += annual_real_approx(row, "pf") * 12
            months += 12
    pre = [row for row in modern if row["date"] <= "2023-11-01"]
    for row in pre:
        if row.get("bancoReal") is not None:
            bank_sum += row["bancoReal"]
        if row.get("pfReal") is not None:
            pf_sum += row["pfReal"]
    comparable = [
        row
        for row in pre
        if row.get("bancoReal") is not None and row.get("pfReal") is not None
    ]
    months += len(comparable)
    fintech = [row["fintechReal"] for row in pre if row.get("fintechReal") is not None]
    return {
        "bancoReal": bank_sum / months,
        "pfReal": pf_sum / months,
        "fintechReal": sum(fintech) / len(fintech),
        "bankMonths": months,
        "pfMonths": months,
        "fintechMonths": len(fintech),
    }


def read_bcra_txt(path: Path, wanted: set[int], daily: bool) -> dict[str, float]:
    """Lee series BCRA y devuelve pesos nominales mensuales.

    Los importes oficiales están expresados en miles de pesos. Para personales
    se suman las dos bandas de plazo mensuales (1936 y 1938). Para plazo fijo se
    agregan por mes ocho series diarias: cuatro estratos para 30–44 días y cuatro
    para 45–59 días.
    """

    totals: dict[str, float] = defaultdict(float)
    with path.open("r", encoding="ascii", errors="ignore", newline="") as handle:
        for line in handle:
            parts = line.strip().split(";")
            if len(parts) != 3:
                continue
            try:
                code = int(parts[0])
            except ValueError:
                continue
            if code not in wanted:
                continue
            date = datetime.strptime(parts[1], "%d/%m/%Y")
            month = date.strftime("%Y-%m")
            if MIRROR_START <= month <= POST_END:
                totals[month] += float(parts[2]) * 1000
    if daily:
        # La suma diaria es el monto efectivamente constituido durante el mes.
        return dict(sorted(totals.items()))
    return dict(sorted(totals.items()))


def read_ipc() -> dict[str, float]:
    path = INDEC_DIR / "serie_ipc_divisiones.csv"
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                rows = list(csv.DictReader(handle, delimiter=";"))
            break
        except UnicodeDecodeError:
            continue
    else:
        raise RuntimeError("No se pudo decodificar el CSV de IPC INDEC")
    result: dict[str, float] = {}
    for row in rows:
        if row["Codigo"] == "0" and row["Region"] == "Nacional":
            period = row["Periodo"]
            if len(period) == 6:
                result[f"{period[:4]}-{period[4:]}"] = float(row["Indice_IPC"].replace(",", "."))
    return result


def read_fintech_stock() -> dict[str, float]:
    """Extrae sólo feb-2026, cuando el stock constante coincide con el nominal.

    El cuadro publica saldos en miles de millones de pesos de febrero de 2026.
    Usar puntos anteriores como si fueran nominales introduciría una segunda
    corrección por IPC. Por eso la exposición se calcula únicamente en feb-2026.
    """

    path = PNFC_DIR / "series-informe-proveedores-no-financieros-credito-junio-2026.xlsx"
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["3"]
    dates = [sheet.cell(5, col).value for col in range(3, 13)]
    values = [sheet.cell(8, col).value for col in range(3, 13)]
    result: dict[str, float] = {}
    for raw_date, raw_value in zip(dates, values):
        if isinstance(raw_date, datetime):
            month = raw_date.strftime("%Y-%m")
        else:
            text = str(raw_date).lower()
            month = "2023-12" if text == "dic-23" else ""
        if month == "2026-02":
            result[month] = float(raw_value) * 1_000_000_000
    workbook.close()
    if "2026-02" not in result:
        raise RuntimeError("No se encontró el stock Fintech de feb-2026 en la fuente PNFC")
    return result


def write_csv(path: Path, rows: Iterable[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def finite(value: Any) -> bool:
    return isinstance(value, (int, float))


def build_records(
    modern: list[dict[str, Any]],
    baselines: dict[str, float | int],
    personal: dict[str, float],
    pf: dict[str, float],
    ipc: dict[str, float],
    fintech_stock: dict[str, float],
) -> list[dict[str, Any]]:
    ipc_ref = ipc[POST_END]
    records: list[dict[str, Any]] = []
    for source in modern:
        month = source["date"][:7]
        if not (MIRROR_START <= month <= POST_END):
            continue
        if month not in ipc:
            raise RuntimeError(f"Falta IPC nacional para {month}")
        factor = ipc_ref / ipc[month]
        banco_real = source.get("bancoReal")
        pf_real = source.get("pfReal")
        fintech_real = source.get("fintechReal")
        banco_gap = banco_real - baselines["bancoReal"] if finite(banco_real) else None
        pf_gap = pf_real - baselines["pfReal"] if finite(pf_real) else None
        fintech_gap = (
            fintech_real - baselines["fintechReal"] if finite(fintech_real) else None
        )
        banco_amount = personal.get(month)
        pf_amount = pf.get(month)
        stock = fintech_stock.get(month)
        banco_nominal = banco_amount * banco_gap / 100 if banco_amount is not None else None
        pf_nominal = pf_amount * pf_gap / 100 if pf_amount is not None else None
        fintech_exposure = (
            stock * fintech_gap / 100 if stock is not None and fintech_gap is not None else None
        )
        records.append(
            {
                "fecha": month,
                "ventana": "post_shock" if POST_START <= month <= POST_END else "espejo",
                "ipc": ipc[month],
                "ipc_ref": ipc_ref,
                "periodo_referencia": POST_END,
                "banco_real": banco_real,
                "banco_promedio_historico": baselines["bancoReal"],
                "banco_brecha_pp": banco_gap,
                "monto_personales_nominal": banco_amount,
                "monto_personales_pesos_constantes": banco_amount * factor if banco_amount is not None else None,
                "impacto_banco_nominal": banco_nominal,
                "impacto_banco_pesos_constantes": banco_nominal * factor if banco_nominal is not None else None,
                "pf_real": pf_real,
                "pf_promedio_historico": baselines["pfReal"],
                "pf_brecha_pp": pf_gap,
                "monto_pf_nominal": pf_amount,
                "monto_pf_pesos_constantes": pf_amount * factor if pf_amount is not None else None,
                "impacto_pf_nominal": pf_nominal,
                "impacto_pf_pesos_constantes": pf_nominal * factor if pf_nominal is not None else None,
                "pf_rendimiento_perdido_pesos_constantes": max(-(pf_nominal * factor), 0) if pf_nominal is not None else None,
                "pf_rendimiento_adicional_pesos_constantes": max(pf_nominal * factor, 0) if pf_nominal is not None else None,
                "fintech_real": fintech_real,
                "fintech_promedio_historico": baselines["fintechReal"],
                "fintech_brecha_pp": fintech_gap,
                "saldo_fintech": stock,
                "exposicion_fintech": fintech_exposure,
                "exposicion_fintech_pesos_constantes": fintech_exposure * factor if fintech_exposure is not None else None,
                "nota": (
                    "Fintech es stock de cartera total del grupo y exposición, no flujo ni intereses cobrados"
                    if stock is not None
                    else ""
                ),
            }
        )
    return records


def rounded_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    cleaned: list[dict[str, Any]] = []
    for row in rows:
        out: dict[str, Any] = {}
        for key, value in row.items():
            if isinstance(value, float):
                out[key] = round(value, 6)
            elif value is None:
                out[key] = ""
            else:
                out[key] = value
        cleaned.append(out)
    return cleaned


def summarise(records: list[dict[str, Any]], baselines: dict[str, float | int]) -> dict[str, Any]:
    post = [row for row in records if row["ventana"] == "post_shock"]
    mirror = [row for row in records if row["ventana"] == "espejo"]

    def total(rows: list[dict[str, Any]], key: str) -> float:
        return sum(float(row[key]) for row in rows if finite(row.get(key)))

    def window(rows: list[dict[str, Any]]) -> dict[str, float | int | str]:
        bank_values = [row["impacto_banco_pesos_constantes"] for row in rows]
        pf_values = [row["impacto_pf_pesos_constantes"] for row in rows]
        bank_net = sum(bank_values)
        pf_net = sum(pf_values)
        return {
            "inicio": rows[0]["fecha"],
            "fin": rows[-1]["fecha"],
            "meses": len(rows),
            "banco_costo_bruto": sum(value for value in bank_values if value > 0),
            "banco_alivio": -sum(value for value in bank_values if value < 0),
            "banco_neto": bank_net,
            "banco_pp_mes": sum(row["banco_brecha_pp"] for row in rows),
            "pf_rendimiento_perdido": -sum(value for value in pf_values if value < 0),
            "pf_rendimiento_adicional": sum(value for value in pf_values if value > 0),
            "pf_neto": pf_net,
            "pf_pp_mes": sum(row["pf_brecha_pp"] for row in rows),
            "pinza_neta_hogar": bank_net - pf_net,
            "monto_personales_constante": total(rows, "monto_personales_pesos_constantes"),
            "monto_pf_constante": total(rows, "monto_pf_pesos_constantes"),
        }

    post_summary = window(post)
    mirror_summary = window(mirror)
    post_summary["banco_por_millon"] = (
        post_summary["banco_neto"] / post_summary["monto_personales_constante"] * 1_000_000
    )
    post_summary["pf_perdida_neta_por_millon"] = (
        -post_summary["pf_neto"] / post_summary["monto_pf_constante"] * 1_000_000
    )
    latest_fintech = next(
        row for row in post if finite(row.get("exposicion_fintech_pesos_constantes"))
    )
    return {
        "referencia": POST_END,
        "baselines": baselines,
        "post": post_summary,
        "mirror": mirror_summary,
        "diferencial_pinza": post_summary["pinza_neta_hogar"] - mirror_summary["pinza_neta_hogar"],
        "fintech": {
            "fecha": latest_fintech["fecha"],
            "saldo_nominal": latest_fintech["saldo_fintech"],
            "brecha_pp": latest_fintech["fintech_brecha_pp"],
            "exposicion_constante": latest_fintech["exposicion_fintech_pesos_constantes"],
        },
    }


def manifest_rows() -> list[dict[str, str]]:
    downloaded = "2026-08-20"
    specs = [
        ("bcra_personales_diario", "prestamos_personales", "BCRA", "Montos operados diarios de préstamos", "https://www.bcra.gob.ar/archivos/Pdfs/PublicacionesEstadisticas/diar_pre.xls", BCRA_DIR / "diar_pre.xls", "", "pre177 / cuenta SISCEN 51", "2021-04/2026-07", "XLS oficial", "Control de conciliación; personales a tasa fija o repactable, total sistema"),
        ("bcra_personales_mensual", "prestamos_personales", "BCRA", "Montos operados mensuales de préstamos", "https://www.bcra.gob.ar/archivos/Pdfs/PublicacionesEstadisticas/preser_mon.xls", BCRA_DIR / "preser_mon.xls", "", "1936 + 1938", "2021-04/2026-07", "XLS oficial", "Columnas de personales hasta 180 días y más de 180 días; miles de pesos"),
        ("bcra_personales_tasas", "prestamos_personales", "BCRA", "Tasas mensuales de préstamos", "https://www.bcra.gob.ar/archivos/Pdfs/PublicacionesEstadisticas/preser_tas.xls", BCRA_DIR / "preser_tas.xls", "", "1937 + 1939 ponderadas por monto", "2002-01/2026-07", "XLS oficial", "Fuente de control de la TNA de personales compatible con los montos"),
        ("bcra_personales_txt", "prestamos_personales", "BCRA", "Series históricas de préstamos al sector privado", "https://www.bcra.gob.ar/archivos/Pdfs/PublicacionesEstadisticas/tas2_ser.txt", BCRA_DIR / "tas2_ser.txt", "", "1936 + 1938", "2021-04/2026-07", "TXT oficial", "Filas mensuales; suma de ambas bandas de plazo; ×1000 para pasar a pesos"),
        ("bcra_pf_historico", "plazo_fijo_30_59", "BCRA", "Tasas y montos de depósitos a plazo fijo", "https://www.bcra.gob.ar/archivos/Pdfs/PublicacionesEstadisticas/pashis.xls", BCRA_DIR / "pashis.xls", "", "pasmes49/51/53/55 + pasmes57/59/61/63", "2021-04/2026-07", "XLS oficial", "Hoja Estra_men; suma de cuatro estratos 30-44 y cuatro estratos 45-59"),
        ("bcra_pf_txt", "plazo_fijo_30_59", "BCRA", "Series históricas de tasas y montos de depósitos", "https://www.bcra.gob.ar/archivos/Pdfs/PublicacionesEstadisticas/tas1_ser.txt", BCRA_DIR / "tas1_ser.txt", "", "1307/1309/1311/1313/1315/1317/1319/1321", "2021-04/2026-07", "TXT oficial", "Filas diarias; suma mensual y ×1000 para pasar de miles de pesos a pesos"),
        ("bcra_boletin_202608", "tasas", "BCRA", "Boletín Estadístico agosto 2026", "https://www.bcra.gob.ar/archivos/Pdfs/PublicacionesEstadisticas/BoletinEstadistico/boldat202608.pdf", BCRA_DIR / "boldat202608.pdf", "", "", "hasta 2026-07", "PDF oficial", "Control de fecha de corte y vínculos a series"),
        ("bcra_metodologia_series", "metodologia", "BCRA", "Metodología de las series estadísticas", "https://www.bcra.gob.ar/Pdfs/PublicacionesEstadisticas/Metodologia_Series.pdf", METHOD_DIR / "Metodologia_Series.pdf", "", "SISCEN 00018 cuenta 51; SISCEN 0002 cuentas 149/152", "vigente a descarga", "PDF oficial", "Define universo, ponderación por montos y exclusiones"),
        ("bcra_catalogo_series", "metodologia", "BCRA", "Catálogo de códigos de series", "https://www.bcra.gob.ar/archivos/Pdfs/PublicacionesEstadisticas/es_series.txt", METHOD_DIR / "es_series.txt", "", "1936/1938 y 1307-1321", "catálogo", "TXT oficial", "Descripción completa de códigos, frecuencia y unidad"),
        ("bcra_info_tasas", "metodologia", "BCRA", "Información de series históricas de tasas", "https://www.bcra.gob.ar/archivos/Pdfs/PublicacionesEstadisticas/es_info_series_tashis.txt", METHOD_DIR / "es_info_series_tashis.txt", "", "", "catálogo", "TXT oficial", "Metadatos complementarios de series"),
        ("indec_ipc", "ipc", "INDEC", "IPC nacional: nivel general y divisiones", "https://www.indec.gob.ar/ftp/cuadros/economia/serie_ipc_divisiones.csv", INDEC_DIR / "serie_ipc_divisiones.csv", "", "Código 0; Región Nacional; Indice_IPC", "2021-04/2026-07", "CSV oficial", "Índice nacional mensual usado para llevar todos los flujos a pesos de julio de 2026"),
        ("indec_ipc_metadatos", "ipc", "INDEC", "Metadatos de las series del IPC", "https://www.indec.gob.ar/ftp/cuadros/economia/serie_ipc_metadatos.txt", INDEC_DIR / "serie_ipc_metadatos.txt", "", "Código 0", "vigente a descarga", "TXT oficial", "Definiciones de variables del CSV de IPC"),
        ("pnfc_informe", "fintech", "BCRA", "Informe de Proveedores No Financieros de Crédito", "https://www.bcra.gob.ar/archivos/Pdfs/PublicacionesEstadisticas/informes/informe-proveedores-no-financieros-credito-junio-2026.pdf", PNFC_DIR / "informe-proveedores-no-financieros-credito-junio-2026.pdf", "2026-06-04", "", "hasta 2026-02", "PDF oficial", "Explica que las tasas se ponderan por saldos y que la publicación trabaja con stocks"),
        ("pnfc_series", "fintech", "BCRA", "Series del informe PNFC junio 2026", "https://www.bcra.gob.ar/archivos/Pdfs/PublicacionesEstadisticas/informes/series-informe-proveedores-no-financieros-credito-junio-2026.xlsx", PNFC_DIR / "series-informe-proveedores-no-financieros-credito-junio-2026.xlsx", "2026-06-04", "Hoja 3, fila Fintech; hoja 5, TNA Fintech", "2026-02", "XLSX oficial", "Stock Fintech total en miles de millones de pesos de feb-2026; se usa sólo como exposición separada"),
        ("pnfc_anexo", "fintech", "BCRA", "Anexo estadístico PNFC junio 2026", "https://www.bcra.gob.ar/archivos/Pdfs/PublicacionesEstadisticas/informes/anexo-estadistico-proveedores-no-financieros-credito-junio-2026.xlsx", PNFC_DIR / "anexo-estadistico-proveedores-no-financieros-credito-junio-2026.xlsx", "2026-06-04", "", "hasta 2026-02", "XLSX oficial", "Control de universos y aperturas por grupo"),
        ("pnfc_pagina", "fintech", "BCRA", "Página de publicación PNFC junio 2026", "https://www.bcra.gob.ar/publicaciones/informe-de-proveedores-no-financieros-de-credito-junio-de-2026/", PNFC_DIR / "informe_pnfc_junio_2026.html", "2026-06-04", "", "hasta 2026-02", "HTML oficial", "Copia local de la página que enlaza PDF, series y anexo"),
    ]
    rows: list[dict[str, str]] = []
    for source_id, topic, institution, title, url, path, publication, code, period, kind, note in specs:
        if not path.exists():
            raise FileNotFoundError(path)
        local = "/data/" + path.relative_to(DATA_DIR).as_posix()
        rows.append(
            {
                "id": source_id,
                "tema": topic,
                "institucion": institution,
                "titulo": title,
                "url_original": url,
                "archivo_local": local,
                "fecha_descarga": downloaded,
                "fecha_publicacion": publication,
                "codigo_serie": code,
                "periodo_utilizado": period,
                "tipo": kind,
                "sha256": sha256(path),
                "nota": note,
            }
        )
    return rows


RATES_PANEL_HTML = r'''
<section class="card" id="ratesMoneySection" style="margin-top:16px">
  <div class="card-head">
    <div class="card-title">④ La pinza financiera en pesos <span>♡</span></div>
    <div class="kicker">costo del crédito y rendimiento del ahorro frente a su propia norma histórica · pesos de jul-2026</div>
  </div>
  <div class="method-legend rates-money-legend">
    <strong>Estado de auditoría:</strong>
    <span class="method-badge official">flujos BCRA oficiales</span>
    <span class="method-badge official">IPC INDEC oficial</span>
    <span class="method-badge derived">impactos derivados</span>
    <span class="method-badge partial">Fintech = stock / exposición</span>
  </div>
  <div class="rates-money-kpi-grid" id="ratesMoneyGrid"></div>
  <div class="rates-money-normalized" id="ratesMoneyNormalized"></div>
  <div id="ratesMoneyChart"></div>
  <div class="rates-fintech-panel">
    <div>
      <div class="tag">Fintech · exposición de cartera al diferencial real</div>
      <div class="big" id="ratesFintechAmount">—</div>
      <div class="mini" id="ratesFintechNote"></div>
    </div>
    <div id="ratesFintechChart"></div>
  </div>
  <div class="rates-money-table-wrap">
    <table class="rates-money-table">
      <thead><tr><th>Mes</th><th>Personales operados</th><th>Brecha banco</th><th>Impacto banco constante</th><th>PF constituidos 30–59</th><th>Brecha PF</th><th>Impacto PF constante</th></tr></thead>
      <tbody id="ratesMoneyTableBody"></tbody>
    </table>
  </div>
  <div class="rates-money-audit-grid">
    <div class="rates-money-audit-card">
      <h3>Qué significa el número</h3>
      <p>Si en un mes se operaron $100.000 millones en préstamos personales y la tasa real quedó 2 puntos porcentuales sobre su norma histórica, el indicador asigna $2.000 millones de costo financiero real adicional a ese flujo, antes de llevarlo a pesos de julio de 2026.</p>
      <p><b>Personales:</b> 1936 + 1938, monto operado mensual a tasa fija o repactable. <b>Plazo fijo:</b> ocho series diarias que cubren cuatro estratos de monto para 30–44 y 45–59 días, agregadas por mes.</p>
    </div>
    <div class="rates-money-audit-card warn">
      <h3>Qué NO significa</h3>
      <ul><li>No es CFT ni incluye todas las comisiones.</li><li>No prueba causalidad individual ni ganancia bancaria.</li><li>No equivale a stock de deuda ni repite un mismo préstamo.</li><li>Los montos operados incluyen refinanciaciones y no son flujo neto.</li><li>Fintech no se suma: es exposición sobre stock, no interés cobrado.</li></ul>
    </div>
  </div>
  <div class="rates-money-actions">
    <button class="download-link" onclick="downloadRatesImpactCsv()" type="button">⬇ CSV · impacto financiero auditado</button>
    <a class="source-link" href="https://www.bcra.gob.ar/archivos/Pdfs/PublicacionesEstadisticas/tas2_ser.txt" rel="noopener" target="_blank">🏦 BCRA · personales · tas2_ser</a>
    <a class="source-link" href="https://www.bcra.gob.ar/archivos/Pdfs/PublicacionesEstadisticas/tas1_ser.txt" rel="noopener" target="_blank">💰 BCRA · PF · tas1_ser</a>
    <a class="source-link" href="https://www.bcra.gob.ar/Pdfs/PublicacionesEstadisticas/Metodologia_Series.pdf" rel="noopener" target="_blank">📚 BCRA · metodología</a>
    <a class="source-link" href="https://www.indec.gob.ar/ftp/cuadros/economia/serie_ipc_divisiones.csv" rel="noopener" target="_blank">📊 INDEC · IPC</a>
    <a class="source-link" href="https://www.bcra.gob.ar/publicaciones/informe-de-proveedores-no-financieros-de-credito-junio-de-2026/" rel="noopener" target="_blank">📱 BCRA · PNFC</a>
  </div>
  <div class="rates-files-audit"><b>Archivos y columnas usados:</b> <code>tas2_ser.txt</code> · 1936 + 1938 · abr-2021→jul-2026; <code>tas1_ser.txt</code> · 1307/1309/1311/1313/1315/1317/1319/1321 · suma diaria→mensual; <code>serie_ipc_divisiones.csv</code> · Código 0 / Nacional / Indice_IPC; <code>series-informe-proveedores-no-financieros-credito-junio-2026.xlsx</code> · hoja 3 / Fintech / feb-2026. Descarga: 20/08/2026. Hashes SHA-256: <code>/data/fuentes/FUENTES.csv</code>.</div>
</section>
'''.strip()


RATES_NOMINAL_CALLOUT_HTML = r'''
          <div class="rates-usury-callout" id="ratesUsuryNominalCallout">
            <div class="rates-usury-callout-values">
              <div class="rates-usury-stat before">
                <div class="eyebrow">Antes · espejo abr-2021→nov-2023</div>
                <div class="amount" id="ratesUsuryBeforeTotal">—</div>
              </div>
              <div class="rates-usury-stat after">
                <div class="eyebrow">Después · dic-2023→jul-2026</div>
                <div class="amount" id="ratesUsuryAfterTotal">—</div>
              </div>
              <div class="rates-usury-stat effect">
                <div class="eyebrow">Efecto neto · después − antes</div>
                <div class="amount" id="ratesUsuryEffectTotal">—</div>
                <div class="mini" id="ratesUsuryEffectPct">—</div>
              </div>
            </div>
            <div class="rates-usury-callout-copy">
              <div class="method-badge-row">
                <span class="method-badge derived">comparación antes / después</span>
                <span class="method-badge official">32 meses por ventana</span>
              </div>
              <b>Efecto mostrado:</b> a cada mes pos-shock le restamos el mes equivalente de la ventana espejo previa y acumulamos la diferencia.
              Las tres curvas usan <code>costo bancario adicional − rendimiento neto del plazo fijo</code> y pesos constantes de julio de 2026.
            </div>
          </div>
'''.strip()


RATES_NOMINAL_PRELUDE = r'''
const ratesAnnualSpreadTna = annual.map(d => d.banco - d.pf);
const ratesUsuryPostRows = __POST_ROWS__;
const ratesUsuryMirrorRows = __MIRROR_ROWS__;
function ratesUsuryMonthClose(ym){
  const [year, month]=ym.split('-').map(Number);
  return new Date(Date.UTC(year,month,0)).toISOString().slice(0,10);
}
function ratesUsuryMonthlyValue(r){
  return Number(r.impacto_banco_pesos_constantes) - Number(r.impacto_pf_pesos_constantes);
}
function ratesUsuryWindow(rows,anchor){
  const x=[decimalYearFromIso(anchor)], y=[0], monthly=[0];
  let running=0;
  rows.forEach(r=>{
    const value=ratesUsuryMonthlyValue(r);
    running+=value;
    x.push(decimalYearFromIso(ratesUsuryMonthClose(r.fecha)));
    y.push(running/1e12);
    monthly.push(value/1e12);
  });
  return {x,y,monthly,total:running};
}
const ratesUsuryMirror=ratesUsuryWindow(ratesUsuryMirrorRows,'2021-03-31');
const ratesUsuryPost=ratesUsuryWindow(ratesUsuryPostRows,'2023-12-10');
const ratesUsuryEffect={
  x:ratesUsuryPost.x,
  y:ratesUsuryPost.y.map((value,index)=>value-(ratesUsuryMirror.y[index]??ratesUsuryMirror.y.at(-1))),
  monthly:ratesUsuryPost.monthly.map((value,index)=>value-(ratesUsuryMirror.monthly[index]??0))
};
function ratesUsuryAxis(mobile){
  const values=[...ratesUsuryMirror.y,...ratesUsuryPost.y,...ratesUsuryEffect.y];
  const low=Math.min(0,...values), high=Math.max(0,...values);
  const pad=Math.max(.35,(high-low)*.08);
  return {
    title:mobile?'$ billones jul-26':'$ billones constantes de jul-2026',
    overlaying:'y',side:'right',range:[low-pad,high+pad],fixedrange:true,
    showgrid:false,zeroline:true,zerolinecolor:'#d7bcc8',tickprefix:'$ ',ticksuffix:' B',
    tickfont:{color:'#8f4665'},titlefont:{color:'#8f4665'}
  };
}
'''.strip()


RATES_NOMINAL_TRACES = r'''
  {
    x:annualPointX,
    y:ratesAnnualSpreadTna,
    customdata:annual.map((d, i)=> d.year===2026 ? 'cierre 2026-07-31 · promedio ene–jul' : `cierre ${annualPointDates[i]} · promedio anual`),
    name:'Diferencial banco − PF (pp TNA)',
    mode:'lines+markers',
    line:{color:'#a978e8',width:2.2,dash:'dot'},
    marker:{size:5,color:'#a978e8'},
    hovertemplate:'Diferencial nominal banco − PF: <b>%{y:.2f} pp TNA</b><br><span style="font-size:11px">%{customdata}</span><extra></extra>'
  },
  {
    x:ratesUsuryMirror.x,
    y:ratesUsuryMirror.y,
    customdata:ratesUsuryMirror.monthly,
    name:'Antes · acumulado espejo ($)',
    mode:'lines+markers',
    yaxis:'y2',
    connectgaps:false,
    line:{color:'#3f8a6c',width:2.7,dash:'dash'},
    marker:{size:3.5,color:'#3f8a6c'},
    hovertemplate:'<b>Antes · acumulado espejo</b>: $ %{y:.2f} billones<br>Aporte neto del mes: $ %{customdata:+.2f} billones<extra></extra>'
  },
  {
    x:ratesUsuryPost.x,
    y:ratesUsuryPost.y,
    customdata:ratesUsuryPost.monthly,
    name:'Después · acumulado pos-shock ($)',
    mode:'lines+markers',
    yaxis:'y2',
    connectgaps:false,
    line:{color:'#9f4f78',width:3,dash:'dashdot'},
    marker:{size:4,color:'#9f4f78'},
    hovertemplate:'<b>Después · acumulado pos-shock</b>: $ %{y:.2f} billones<br>Aporte neto del mes: $ %{customdata:+.2f} billones<extra></extra>'
  },
  {
    x:ratesUsuryEffect.x,
    y:ratesUsuryEffect.y,
    customdata:ratesUsuryEffect.monthly,
    name:'Efecto · después − antes ($)',
    mode:'lines+markers',
    yaxis:'y2',
    connectgaps:false,
    line:{color:'#5b3fa3',width:3.4},
    marker:{size:4.5,color:'#5b3fa3'},
    hovertemplate:'<b>Efecto acumulado vs antes</b>: $ %{y:.2f} billones<br>Diferencia del mes: $ %{customdata:+.2f} billones<extra></extra>'
  }
'''.strip()


MILEI_AUDIT_HTML = r'''
    <section class="milei-cost-audit milei-financial-audit" id="mileiFinancialAuditCard">
      <h3>Pinza financiera · cálculo en auditoría</h3>
      <div id="mileiFinancialAuditContent"></div>
      <div class="jump"><button class="subbtn" type="button" onclick="activateTabAndScroll('tab-rates','ratesMoneySection')">Ver tasas, flujos y fuentes →</button></div>
    </section>
'''.rstrip()


RATES_CSS = r'''
<style id="rates-money-audited-v122">
#nominalChart{height:650px}
.rates-usury-callout{display:grid;grid-template-columns:minmax(340px,1fr) minmax(260px,.8fr);gap:15px;align-items:center;margin:-2px 20px 12px;padding:14px 16px;border:1px solid #d9cce8;border-radius:18px;background:linear-gradient(135deg,#fbfffc,#fff7fb 58%,#faf7ff);box-sizing:border-box}
.rates-usury-callout-values{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:9px;min-width:0}.rates-usury-stat{min-width:0;padding:10px;border:1px solid #e9dfed;border-radius:14px;background:#fff;box-sizing:border-box}.rates-usury-stat.before{border-color:#cce5d8;background:#f8fffb}.rates-usury-stat.after{border-color:#e8cad7;background:#fff8fb}.rates-usury-stat.effect{border-color:#d9cff1;background:#fbf9ff}.rates-usury-callout .eyebrow{font-size:8.5px;font-weight:900;text-transform:uppercase;letter-spacing:.025em;color:#846f8b;line-height:1.35}.rates-usury-callout .amount{margin-top:5px;font-size:21px;line-height:1;font-weight:950;color:#715080;white-space:nowrap}.rates-usury-stat.before .amount{color:#3f8a6c}.rates-usury-stat.after .amount{color:#9f4f78}.rates-usury-stat.effect .amount{color:#5b3fa3}.rates-usury-stat .mini{margin-top:5px;font-size:9px;font-weight:850;color:#76667e}.rates-usury-callout-copy{font-size:10.5px;line-height:1.55;color:#715f76}.rates-usury-callout-copy b{color:#5b3fa3}.rates-usury-callout-copy code{font-size:9.5px;color:#694c75;overflow-wrap:anywhere}.rates-usury-callout .method-badge-row{margin:0 0 7px}
#ratesMoneySection{overflow:visible}
#ratesMoneySection .rates-money-legend{margin:0 20px 16px}
.rates-money-kpi-grid{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:12px;margin:0 20px 16px}
.rates-money-kpi{min-width:0;padding:15px;border:1px solid #e7d8e8;border-radius:18px;background:#fff;box-sizing:border-box}
.rates-money-kpi.bank{border-color:#c8d8ff;background:linear-gradient(180deg,#f7f9ff,#fff)}
.rates-money-kpi.pf{border-color:#ffd4bd;background:linear-gradient(180deg,#fff9f5,#fff)}
.rates-money-kpi.diff{border-color:#e9bfd1;background:linear-gradient(180deg,#fff7fa,#fff)}
.rates-money-kpi .tag,.rates-fintech-panel .tag{font-size:9px;font-weight:900;text-transform:uppercase;letter-spacing:.035em;color:#8d718f}
.rates-money-kpi .big,.rates-fintech-panel .big{margin-top:7px;font-size:25px;line-height:1.02;font-weight:950;color:#724862}
.rates-money-kpi.bank .big{color:#3669c9}.rates-money-kpi.pf .big{color:#ba6538}.rates-money-kpi.diff .big{color:#a54569}
.rates-money-kpi .mini,.rates-fintech-panel .mini{margin-top:7px;font-size:10px;line-height:1.5;color:#756579}
.rates-money-normalized{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:11px;margin:0 20px 12px}
.rates-money-normalized>div{padding:12px 14px;border:1px dashed #d9c7df;border-radius:15px;background:#fcf9ff;font-size:11px;line-height:1.5;color:#705c78;box-sizing:border-box}
#ratesMoneyChart{min-height:430px;margin:0 20px 6px}
.rates-fintech-panel{display:grid;grid-template-columns:minmax(260px,.8fr) minmax(0,1.2fr);gap:16px;align-items:center;margin:12px 20px;padding:15px;border:1px solid #edcbd9;border-radius:18px;background:#fff8fb;box-sizing:border-box}
#ratesFintechChart{min-height:210px}
.rates-money-table-wrap{margin:16px 20px;overflow-x:auto;border:1px solid #e6dbe9;border-radius:16px;background:#fff}
.rates-money-table{width:100%;min-width:1040px;border-collapse:collapse;font-size:10px}
.rates-money-table th{position:sticky;top:0;padding:10px;text-align:left;background:#f7f2fa;color:#80698e;font-size:9px;text-transform:uppercase;letter-spacing:.025em;border-bottom:1px solid #e6dce9}
.rates-money-table td{padding:9px 10px;border-bottom:1px solid #f1eaf3;color:#705d76;white-space:nowrap}
.rates-money-table tr:last-child td{border-bottom:0}.rates-money-table .pos{color:#a34568;font-weight:850}.rates-money-table .neg{color:#43866b;font-weight:850}
.rates-money-audit-grid{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:12px;margin:16px 20px}
.rates-money-audit-card{padding:15px;border:1px solid #d7e3ff;border-radius:17px;background:#f8faff;box-sizing:border-box}
.rates-money-audit-card.warn{border-color:#f0d7a5;background:#fffdf6}
.rates-money-audit-card h3{margin:0 0 8px;font-size:15px;color:#644b70}.rates-money-audit-card p,.rates-money-audit-card li{font-size:10.5px;line-height:1.55;color:#746479}.rates-money-audit-card p:last-child{margin-bottom:0}.rates-money-audit-card ul{margin:0;padding-left:18px}
#ratesMoneySection .rates-money-actions{margin:14px 20px 0}.rates-files-audit{margin:13px 20px 20px;padding:12px;border-radius:14px;background:#f8f3fb;font-size:9.5px;line-height:1.55;color:#735f7c;overflow-wrap:anywhere}
#tab-milei-cost .milei-financial-audit{border-color:#e8c6d5;background:linear-gradient(180deg,#fff8fb,#fff)}
#tab-milei-cost .milei-financial-audit .audit-amount{font-size:28px;line-height:1;font-weight:950;color:#a54569;margin:7px 0}
#tab-milei-cost .milei-financial-audit p{font-size:11px;line-height:1.55;color:#756278;margin:5px 0}
@media(max-width:1366px){.rates-money-kpi-grid{grid-template-columns:repeat(2,minmax(0,1fr))}}
@media(max-width:1024px){.rates-usury-callout{grid-template-columns:1fr}.rates-fintech-panel{grid-template-columns:1fr}.rates-money-audit-grid{grid-template-columns:1fr}}
@media(max-width:768px){#ratesMoneyChart{min-height:500px}.rates-money-normalized{grid-template-columns:1fr}}
@media(max-width:720px){#nominalChart{height:720px}.rates-usury-callout{grid-template-columns:1fr;margin-left:14px;margin-right:14px}.rates-usury-callout-values{grid-template-columns:1fr}.rates-usury-callout .amount{font-size:21px}#ratesMoneySection .rates-money-legend,.rates-money-kpi-grid,.rates-money-normalized,#ratesMoneyChart,.rates-fintech-panel,.rates-money-table-wrap,.rates-money-audit-grid,#ratesMoneySection .rates-money-actions,.rates-files-audit{margin-left:14px;margin-right:14px}.rates-money-kpi-grid{grid-template-columns:1fr}.rates-money-kpi .big{font-size:23px}.rates-fintech-panel{padding:13px}#ratesFintechChart{min-height:190px}}
@media(max-width:430px){#nominalChart{height:740px}.rates-usury-callout{margin-left:11px;margin-right:11px;padding:12px}.rates-usury-callout .amount{font-size:20px}#ratesMoneySection .rates-money-legend,.rates-money-kpi-grid,.rates-money-normalized,#ratesMoneyChart,.rates-fintech-panel,.rates-money-table-wrap,.rates-money-audit-grid,#ratesMoneySection .rates-money-actions,.rates-files-audit{margin-left:11px;margin-right:11px}.rates-money-kpi{padding:13px}.rates-money-kpi .big{font-size:21px}#ratesMoneyChart{min-height:540px}}
@media(max-width:390px){#ratesMoneySection .rates-money-legend,.rates-money-kpi-grid,.rates-money-normalized,#ratesMoneyChart,.rates-fintech-panel,.rates-money-table-wrap,.rates-money-audit-grid,#ratesMoneySection .rates-money-actions,.rates-files-audit{margin-left:9px;margin-right:9px}}
</style>
'''.strip()


RATES_JS = r'''
const ratesMoneySummary = __SUMMARY__;
const ratesMoneyRows = __ROWS__;
const ratesImpactCsv = __CSV__;

function ratesMoneyArs(value, digits=2){
  const abs=Math.abs(Number(value));
  const sign=value<0?'−':'';
  if(abs>=1e12)return `${sign}$ ${(abs/1e12).toLocaleString('es-AR',{minimumFractionDigits:digits,maximumFractionDigits:digits})} billones`;
  if(abs>=1e9)return `${sign}$ ${(abs/1e9).toLocaleString('es-AR',{minimumFractionDigits:1,maximumFractionDigits:1})} mil M`;
  return `${sign}$ ${abs.toLocaleString('es-AR',{maximumFractionDigits:0})}`;
}
function ratesMoneyCompact(value){const v=Number(value);return `${v<0?'−':''}$ ${(Math.abs(v)/1e12).toLocaleString('es-AR',{minimumFractionDigits:2,maximumFractionDigits:2})} B`;}
function ratesMoneySignedClass(value){return Number(value)>=0?'pos':'neg';}
function renderRatesMoney(){
  const s=ratesMoneySummary,p=s.post,m=s.mirror;
  const el=document.getElementById('ratesMoneyGrid');
  if(!el)return;
  el.innerHTML=`
    <div class="rates-money-kpi bank"><div class="tag">Préstamos personales · pos-shock</div><div class="big">${ratesMoneyArs(p.banco_neto)}</div><div class="mini">Saldo neto de costo financiero real adicional. Costo bruto: <b>${ratesMoneyArs(p.banco_costo_bruto)}</b> · alivio: <b>${ratesMoneyArs(p.banco_alivio)}</b>.</div></div>
    <div class="rates-money-kpi pf"><div class="tag">Plazo fijo · pos-shock</div><div class="big">${ratesMoneyArs(p.pf_rendimiento_perdido)}</div><div class="mini">Rendimiento real dejado de percibir en meses peores a la media. Meses mejores: <b>${ratesMoneyArs(p.pf_rendimiento_adicional)}</b> · saldo firmado: <b>${ratesMoneyArs(p.pf_neto)}</b>.</div></div>
    <div class="rates-money-kpi"><div class="tag">Ventana espejo · ${m.inicio}→${m.fin}</div><div class="big">${ratesMoneyArs(m.pinza_neta_hogar)}</div><div class="mini">Resultado comparable de ${m.meses} meses: costo bancario neto menos rendimiento neto del plazo fijo.</div></div>
    <div class="rates-money-kpi diff"><div class="tag">Diferencial pos-shock vs espejo</div><div class="big">${ratesMoneyArs(s.diferencial_pinza)}</div><div class="mini">Cambio de la pinza neta para el hogar entre ventanas iguales, todo a pesos de ${s.referencia}.</div></div>`;
  const beforeTotal=document.getElementById('ratesUsuryBeforeTotal');
  const afterTotal=document.getElementById('ratesUsuryAfterTotal');
  const effectTotal=document.getElementById('ratesUsuryEffectTotal');
  const effectPct=document.getElementById('ratesUsuryEffectPct');
  if(beforeTotal)beforeTotal.textContent=ratesMoneyCompact(m.pinza_neta_hogar);
  if(afterTotal)afterTotal.textContent=ratesMoneyCompact(p.pinza_neta_hogar);
  if(effectTotal)effectTotal.textContent=ratesMoneyCompact(s.diferencial_pinza);
  if(effectPct)effectPct.textContent=`${((p.pinza_neta_hogar/m.pinza_neta_hogar-1)*100).toLocaleString('es-AR',{minimumFractionDigits:1,maximumFractionDigits:1})}% vs antes`;
  document.getElementById('ratesMoneyNormalized').innerHTML=`<div><b>Por cada $1 millón operado en personales:</b> ${ratesMoneyArs(p.banco_por_millon,0)} de diferencial neto ponderado por los flujos reales.</div><div><b>Por cada $1 millón constituido en PF 30–59:</b> ${ratesMoneyArs(p.pf_perdida_neta_por_millon,0)} de pérdida neta ponderada para el ahorrista.</div>`;
  const table=document.getElementById('ratesMoneyTableBody');
  table.innerHTML=ratesMoneyRows.map(r=>`<tr><td>${r.fecha}</td><td>${ratesMoneyArs(r.monto_personales_nominal)}</td><td class="${ratesMoneySignedClass(r.banco_brecha_pp)}">${Number(r.banco_brecha_pp).toLocaleString('es-AR',{minimumFractionDigits:2,maximumFractionDigits:2})} pp</td><td class="${ratesMoneySignedClass(r.impacto_banco_pesos_constantes)}">${ratesMoneyArs(r.impacto_banco_pesos_constantes)}</td><td>${ratesMoneyArs(r.monto_pf_nominal)}</td><td class="${ratesMoneySignedClass(r.pf_brecha_pp)}">${Number(r.pf_brecha_pp).toLocaleString('es-AR',{minimumFractionDigits:2,maximumFractionDigits:2})} pp</td><td class="${ratesMoneySignedClass(r.impacto_pf_pesos_constantes)}">${ratesMoneyArs(r.impacto_pf_pesos_constantes)}</td></tr>`).join('');
  const fintech=s.fintech;
  document.getElementById('ratesFintechAmount').textContent=ratesMoneyArs(fintech.exposicion_constante);
  document.getElementById('ratesFintechNote').innerHTML=`Foto de <b>${fintech.fecha}</b>: stock Fintech total ${ratesMoneyArs(fintech.saldo_nominal)} × brecha real de ${Number(fintech.brecha_pp).toLocaleString('es-AR',{minimumFractionDigits:2,maximumFractionDigits:2})} pp, reexpresado a ${s.referencia}. No equivale a intereses cobrados ni a crédito nuevo.`;
  const milei=document.getElementById('mileiFinancialAuditContent');
  if(milei)milei.innerHTML=`<div class="audit-amount">${ratesMoneyArs(p.pinza_neta_hogar)}</div><p>Pinza neta pos-shock estimada para los flujos bancarios auditados, a pesos de ${s.referencia}. Diferencial contra la ventana espejo: <b>${ratesMoneyArs(s.diferencial_pinza)}</b>.</p><p><b>No se suma al total del tab:</b> permanece como tarjeta separada hasta descartar doble conteo con otros contrafactuales.</p>`;
  renderRatesMoneyCharts();
}
function renderRatesMoneyCharts(){
  if(typeof Plotly==='undefined')return;
  const s=ratesMoneySummary,p=s.post,m=s.mirror,mobile=window.innerWidth<=720;
  Plotly.react('ratesMoneyChart',[{type:'bar',orientation:'h',x:[p.banco_neto,m.banco_neto,-p.pf_neto,-m.pf_neto],y:['Banco · pos-shock','Banco · espejo','PF · pos-shock','PF · espejo'],marker:{color:['#4d83ff','#9bb6f7','#ff8d52','#ffc3a4']},text:[p.banco_neto,m.banco_neto,-p.pf_neto,-m.pf_neto].map(ratesMoneyCompact),textposition:'outside',cliponaxis:false,hovertemplate:'<b>%{y}</b><br>Impacto neto hogar: <b>$%{x:,.0f}</b><extra></extra>'}],{paper_bgcolor:'rgba(255,255,255,0)',plot_bgcolor:'#fffdfd',font:{color:'#5e4670',family:'Inter,system-ui,sans-serif'},margin:{l:mobile?128:170,r:mobile?58:115,t:50,b:64},xaxis:{title:'Pesos constantes de jul-2026 · positivo = costo o pérdida para el hogar',gridcolor:'#efe4f4',zeroline:true,zerolinecolor:'#bba9c4',fixedrange:true},yaxis:{automargin:true,fixedrange:true},showlegend:false},{responsive:true,displaylogo:false,displayModeBar:false,scrollZoom:false,doubleClick:false});
  Plotly.react('ratesFintechChart',[{type:'bar',orientation:'h',x:[s.fintech.exposicion_constante],y:['Fintech · feb-2026'],marker:{color:'#ff6387'},text:[ratesMoneyCompact(s.fintech.exposicion_constante)],textposition:'outside',cliponaxis:false,hovertemplate:'Exposición a la brecha real: <b>$%{x:,.0f}</b><extra></extra>'}],{paper_bgcolor:'rgba(255,255,255,0)',plot_bgcolor:'#fff8fb',font:{color:'#5e4670',family:'Inter,system-ui,sans-serif'},margin:{l:mobile?120:145,r:mobile?55:90,t:25,b:48},xaxis:{title:'Pesos constantes de jul-2026',gridcolor:'#f0dce5',fixedrange:true},yaxis:{fixedrange:true},showlegend:false},{responsive:true,displaylogo:false,displayModeBar:false,scrollZoom:false,doubleClick:false});
}
function downloadRatesImpactCsv(){
  const blob=new Blob(['\ufeff'+ratesImpactCsv],{type:'text/csv;charset=utf-8'});
  const url=URL.createObjectURL(blob),a=document.createElement('a');
  a.href=url;a.download='tasas_impacto_financiero_post_shock.csv';document.body.appendChild(a);a.click();a.remove();setTimeout(()=>URL.revokeObjectURL(url),500);
}
renderRatesMoney();
'''.strip()


def build_html(
    source_html: str,
    summary: dict[str, Any],
    post_rows: list[dict[str, Any]],
    mirror_rows: list[dict[str, Any]],
    csv_content: str,
) -> str:
    html = source_html
    panel_pattern = re.compile(
        r'<section class="card" id="ratesMoneySection".*?</section>\s*(?=<div class="sources-box">)',
        re.S,
    )
    html, count = panel_pattern.subn(RATES_PANEL_HTML + "\n\n    ", html, count=1)
    if count != 1:
        raise RuntimeError("No se pudo reemplazar el panel monetario v119")

    nominal_chart = '          <div id="nominalChart"></div>'
    if nominal_chart not in html:
        raise RuntimeError("No se encontró el gráfico nominal para insertar el saldo editorial")
    html = html.replace(
        nominal_chart,
        nominal_chart + "\n" + RATES_NOMINAL_CALLOUT_HTML,
        1,
    )
    nominal_kicker = (
        '<div class="kicker">inflación acumulada + TNA promedio · '
        '2002–2026 · 2026 parcial</div>'
    )
    if nominal_kicker not in html:
        raise RuntimeError("No se encontró el subtítulo del gráfico nominal")
    html = html.replace(
        nominal_kicker,
        '<div class="kicker">inflación + TNA + diferencial + efecto monetario antes/después · '
        '2002–2026 · 2026 parcial</div>',
        1,
    )

    def compact_impact_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
        return [
            {
                "fecha": row["fecha"],
                "impacto_banco_pesos_constantes": row[
                    "impacto_banco_pesos_constantes"
                ],
                "impacto_pf_pesos_constantes": row[
                    "impacto_pf_pesos_constantes"
                ],
            }
            for row in rows
        ]

    nominal_prelude = (
        RATES_NOMINAL_PRELUDE.replace(
            "__POST_ROWS__",
            json.dumps(
                compact_impact_rows(post_rows),
                ensure_ascii=False,
                separators=(",", ":"),
            ),
        ).replace(
            "__MIRROR_ROWS__",
            json.dumps(
                compact_impact_rows(mirror_rows),
                ensure_ascii=False,
                separators=(",", ":"),
            ),
        )
    )
    nominal_needle = "const nominalTraces = ["
    if nominal_needle not in html:
        raise RuntimeError("No se encontró nominalTraces")
    html = html.replace(
        nominal_needle,
        nominal_prelude + "\n\n" + nominal_needle,
        1,
    )
    nominal_traces_pattern = re.compile(
        r"(const nominalTraces = \[.*?)(\n\];\n\nconst annualMandateX)",
        re.S,
    )
    html, count = nominal_traces_pattern.subn(
        lambda match: match.group(1) + ",\n" + RATES_NOMINAL_TRACES + match.group(2),
        html,
        count=1,
    )
    if count != 1:
        raise RuntimeError("No se pudieron agregar las líneas al gráfico nominal")

    initial_nominal_axis = (
        "  yaxis:{...common.yaxis,title:'% · inflación acumulada / TNA promedio'},\n"
        "  shapes: nominalDesktopDecor.shapes,"
    )
    if initial_nominal_axis not in html:
        raise RuntimeError("No se encontró el eje del gráfico nominal inicial")
    html = html.replace(
        initial_nominal_axis,
        "  yaxis:{...common.yaxis,title:'% · inflación acumulada / TNA promedio'},\n"
        "  yaxis2:ratesUsuryAxis(false),\n"
        "  shapes: nominalDesktopDecor.shapes,",
        1,
    )
    initial_nominal_annotations = "  annotations: nominalDesktopDecor.annotations\n"
    if initial_nominal_annotations not in html:
        raise RuntimeError("No se encontró el margen del gráfico nominal inicial")
    html = html.replace(
        initial_nominal_annotations,
        "  annotations: nominalDesktopDecor.annotations,\n"
        "  margin:{l:72,r:82,t:205,b:64}\n",
        1,
    )

    visible_nominal_axis = (
        "    yaxis:{\n"
        "      ...common.yaxis,\n"
        "      title:'% · inflación acumulada / TNA promedio',\n"
        "      range:[0,230],\n"
        "      fixedrange:true\n"
        "    },\n"
        "    shapes: nominalDecor.shapes,"
    )
    if visible_nominal_axis not in html:
        raise RuntimeError("No se encontró el eje nominal responsive")
    html = html.replace(
        visible_nominal_axis,
        "    yaxis:{\n"
        "      ...common.yaxis,\n"
        "      title:'% · inflación acumulada / TNA promedio',\n"
        "      range:[0,230],\n"
        "      fixedrange:true\n"
        "    },\n"
        "    yaxis2:ratesUsuryAxis(mobile),\n"
        "    shapes: nominalDecor.shapes,",
        1,
    )
    visible_nominal_margin = (
        "    margin:{l:mobile?55:72,r:mobile?12:24,t:mobile?180:142,b:mobile?70:64}"
    )
    if visible_nominal_margin not in html:
        raise RuntimeError("No se encontró el margen nominal responsive")
    html = html.replace(
        visible_nominal_margin,
        "    margin:{l:mobile?55:72,r:mobile?94:82,t:mobile?225:205,b:mobile?70:64}",
        1,
    )
    responsive_nominal_margins = (
        "      'margin.l': mobile ? 55 : 72,\n"
        "      'margin.r': mobile ? 8 : 30,\n"
        "      'margin.t': mobile ? 172 : 118,\n"
        "      'margin.b': mobile ? 72 : 72,"
    )
    if responsive_nominal_margins not in html:
        raise RuntimeError("No se encontró el ajuste responsive global del gráfico nominal")
    html = html.replace(
        responsive_nominal_margins,
        "      'margin.l': mobile ? 55 : 72,\n"
        "      'margin.r': mobile ? 94 : 82,\n"
        "      'margin.t': mobile ? 220 : 205,\n"
        "      'margin.b': mobile ? 72 : 72,",
        1,
    )
    nominal_responsive_legend = (
        "    ['nominalChart', {\n"
        "      ...mobileLegend,\n"
        "      'margin.l': mobile ? 55 : 72,"
    )
    if nominal_responsive_legend not in html:
        raise RuntimeError("No se encontró la leyenda responsive del gráfico nominal")
    html = html.replace(
        nominal_responsive_legend,
        "    ['nominalChart', {\n"
        "      ...mobileLegend,\n"
        "      'legend.y': mobile ? 1.13 : 1.14,\n"
        "      'margin.l': mobile ? 55 : 72,",
        1,
    )
    html = html.replace("</head>", RATES_CSS + "\n</head>", 1)
    needle = '    <div class="milei-cost-grid" id="mileiCostCards"></div>'
    if needle not in html:
        raise RuntimeError("No se encontró el punto de inserción del tab Milei")
    html = html.replace(needle, needle + "\n\n" + MILEI_AUDIT_HTML, 1)
    js = (
        RATES_JS.replace("__SUMMARY__", json.dumps(summary, ensure_ascii=False, separators=(",", ":")))
        .replace("__ROWS__", json.dumps(post_rows, ensure_ascii=False, separators=(",", ":")))
        .replace("__CSV__", json.dumps(csv_content, ensure_ascii=False))
    )
    js_pattern = re.compile(
        r"function ratesMoneyFmt\(v\).*?renderRatesMoney\(\);",
        re.S,
    )
    html, count = js_pattern.subn(lambda _: js, html, count=1)
    if count != 1:
        raise RuntimeError("No se pudo reemplazar el JavaScript monetario v119")
    return html


def main() -> None:
    DERIVED_DIR.mkdir(parents=True, exist_ok=True)
    original_hash = sha256(INPUT_HTML)
    source_html = INPUT_HTML.read_text(encoding="utf-8")
    annual = js_json_array(source_html, "annual")
    modern = js_json_array(source_html, "modern")
    baselines = historical_baselines(annual, modern)

    personal = read_bcra_txt(BCRA_DIR / "tas2_ser.txt", PERSONAL_CODES, daily=False)
    pf = read_bcra_txt(BCRA_DIR / "tas1_ser.txt", PF_CODES, daily=True)
    ipc = read_ipc()
    fintech_stock = read_fintech_stock()

    required_months = {
        f"{year:04d}-{month:02d}"
        for year in range(2021, 2027)
        for month in range(1, 13)
        if MIRROR_START <= f"{year:04d}-{month:02d}" <= POST_END
    }
    for name, series in (("personales", personal), ("plazo fijo", pf), ("IPC", ipc)):
        missing = sorted(required_months - set(series))
        if missing:
            raise RuntimeError(f"Faltan meses en {name}: {missing}")

    personal_rows = [
        {
            "fecha": month,
            "monto_operado_personales_nominal": round(value, 2),
            "fuente": "BCRA tas2_ser.txt",
            "codigo_serie": "1936+1938",
            "nota_metodologica": "Suma mensual de personales hasta 180 días y más de 180 días; miles de pesos ×1000; incluye nuevas operaciones y refinanciaciones, no es flujo neto",
        }
        for month, value in personal.items()
    ]
    pf_rows = [
        {
            "fecha": month,
            "monto_constituido_pf_30_59_nominal": round(value, 2),
            "fuente": "BCRA tas1_ser.txt",
            "codigo_serie": "1307+1309+1311+1313+1315+1317+1319+1321",
            "nota_metodologica": "Suma mensual de montos diarios, cuatro estratos 30-44 días + cuatro estratos 45-59 días; miles de pesos ×1000",
        }
        for month, value in pf.items()
    ]
    fintech_rows = [
        {
            "fecha": month,
            "saldo_fintech_nominal": round(value, 2),
            "fuente": "BCRA series PNFC junio 2026.xlsx",
            "hoja_fila": "Hoja 3; fila Fintech; columna feb-2026",
            "clasificacion": "stock/proxy",
            "nota_metodologica": "Stock total del grupo Fintech; sólo feb-2026 porque la serie histórica está publicada a pesos constantes de ese mes; no es flujo ni cartera exclusivamente personal",
        }
        for month, value in fintech_stock.items()
    ]
    write_csv(
        DERIVED_DIR / "prestamos_personales_monto_operado_mensual.csv",
        personal_rows,
        list(personal_rows[0]),
    )
    write_csv(
        DERIVED_DIR / "plazo_fijo_30_59_monto_mensual.csv",
        pf_rows,
        list(pf_rows[0]),
    )
    write_csv(
        DERIVED_DIR / "fintech_saldos_mensuales.csv", fintech_rows, list(fintech_rows[0])
    )

    records = build_records(modern, baselines, personal, pf, ipc, fintech_stock)
    summary = summarise(records, baselines)
    fields = list(records[0])
    cleaned = rounded_rows(records)
    post_cleaned = [row for row in cleaned if row["ventana"] == "post_shock"]
    mirror_cleaned = [row for row in cleaned if row["ventana"] == "espejo"]
    accumulated_post = sum(
        row["impacto_banco_pesos_constantes"]
        - row["impacto_pf_pesos_constantes"]
        for row in post_cleaned
    )
    accumulated_mirror = sum(
        row["impacto_banco_pesos_constantes"]
        - row["impacto_pf_pesos_constantes"]
        for row in mirror_cleaned
    )
    if abs(accumulated_post - summary["post"]["pinza_neta_hogar"]) > 1:
        raise RuntimeError("El acumulado pos-shock no coincide con la pinza neta")
    if abs(accumulated_mirror - summary["mirror"]["pinza_neta_hogar"]) > 1:
        raise RuntimeError("El acumulado espejo no coincide con la pinza neta previa")
    if (
        abs(
            (accumulated_post - accumulated_mirror)
            - summary["diferencial_pinza"]
        )
        > 1
    ):
        raise RuntimeError("El efecto antes/después no coincide con el resumen")
    post_csv_path = DERIVED_DIR / "tasas_impacto_financiero_post_shock.csv"
    mirror_csv_path = DERIVED_DIR / "tasas_impacto_financiero_ventana_espejo.csv"
    write_csv(post_csv_path, post_cleaned, fields)
    write_csv(mirror_csv_path, mirror_cleaned, fields)
    (DERIVED_DIR / "tasas_resumen_auditado.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    manifest = manifest_rows()
    manifest_fields = [
        "id",
        "tema",
        "institucion",
        "titulo",
        "url_original",
        "archivo_local",
        "fecha_descarga",
        "fecha_publicacion",
        "codigo_serie",
        "periodo_utilizado",
        "tipo",
        "sha256",
        "nota",
    ]
    write_csv(SOURCE_DIR / "FUENTES.csv", manifest, manifest_fields)

    csv_content = post_csv_path.read_text(encoding="utf-8-sig")
    html = build_html(
        source_html,
        summary,
        post_cleaned,
        mirror_cleaned,
        csv_content,
    )
    OUTPUT_HTML.write_text(html, encoding="utf-8", newline="")
    ROOT_OUTPUT_HTML.write_text(html, encoding="utf-8", newline="")

    if sha256(INPUT_HTML) != original_hash:
        raise RuntimeError("El v119 cambió durante la construcción")
    if rates_money_csv_from_html(html) != csv_content:
        raise RuntimeError("El CSV embebido no coincide con el derivado")
    if sha256(OUTPUT_HTML) != sha256(ROOT_OUTPUT_HTML):
        raise RuntimeError("La copia raíz del v122 no coincide con la versión de data")
    print(
        json.dumps(
            {
                "output_data": str(OUTPUT_HTML),
                "output_root": str(ROOT_OUTPUT_HTML),
                "v119_sha256": original_hash,
                "summary": summary,
            },
            ensure_ascii=False,
            indent=2,
        )
    )


def rates_money_csv_from_html(html: str) -> str:
    match = re.search(r"const ratesImpactCsv = (.*?);\n", html)
    if not match:
        raise RuntimeError("No se encontró el CSV embebido en v122")
    return json.loads(match.group(1))


if __name__ == "__main__":
    main()
