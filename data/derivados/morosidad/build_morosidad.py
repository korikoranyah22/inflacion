#!/usr/bin/env python3
"""Construye el tab reproducible de morosidad sobre el dashboard más reciente.

Fuentes numéricas principales:
- Informe sobre Bancos, mayo de 2026 (BCRA), hoja 9.
- Informe PNFC, junio de 2026 (BCRA), hojas 6 y 7.
- Informe de Inclusión Financiera, abril de 2026 (BCRA), gráfico 2.3.5
  y texto de la página 26.

El script no interpola observaciones ni mezcla porcentajes de saldos con
porcentajes de personas. La opción --input permite fijar el dashboard de base.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import hashlib
import json
import math
import re
import statistics
from pathlib import Path

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover - mensaje operativo
    raise SystemExit(
        "Falta openpyxl. Ejecutar con el Python incluido en el runtime de Codex."
    ) from exc


ROOT = Path(__file__).resolve().parents[3]
DATA = ROOT / "data"
DERIVED = DATA / "derivados" / "morosidad"
SOURCES = DATA / "fuentes" / "morosidad"


def fmt(value: float, digits: int = 2) -> str:
    return f"{value:.{digits}f}".replace(".", ",")


def iso_month(value) -> str | None:
    if isinstance(value, (dt.datetime, dt.date)):
        return value.strftime("%Y-%m-01")
    if isinstance(value, str) and "-" in value:
        months = {
            "ene": 1, "feb": 2, "mar": 3, "abr": 4, "may": 5, "jun": 6,
            "jul": 7, "ago": 8, "sep": 9, "oct": 10, "nov": 11, "dic": 12,
        }
        name, year = value.strip().lower().split("-", 1)
        if name[:3] in months:
            return f"20{int(year):02d}-{months[name[:3]]:02d}-01"
    return None


def add_month(date: str, delta: int) -> str:
    year, month = int(date[:4]), int(date[5:7])
    idx = year * 12 + month - 1 + delta
    return f"{idx // 12:04d}-{idx % 12 + 1:02d}-01"


def pearson(xs: list[float], ys: list[float]) -> float:
    mx, my = statistics.fmean(xs), statistics.fmean(ys)
    numerator = sum((x - mx) * (y - my) for x, y in zip(xs, ys))
    denominator = math.sqrt(
        sum((x - mx) ** 2 for x in xs) * sum((y - my) ** 2 for y in ys)
    )
    return numerator / denominator if denominator else float("nan")


def write_csv(path: Path, rows: list[dict], fieldnames: list[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        raise ValueError(f"No hay filas para {path.name}")
    fields = fieldnames or list(rows[0])
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def read_csv_text(path: Path) -> str:
    return path.read_text(encoding="utf-8-sig")


def latest_dashboard() -> Path:
    candidates = []
    for path in DATA.glob("dashboard_kawaii_*.html"):
        match = re.search(r"dashboard_kawaii_(\d+)", path.name)
        if match:
            candidates.append((int(match.group(1)), path))
    if not candidates:
        raise FileNotFoundError("No se encontró /data/dashboard_kawaii_*.html")
    return max(candidates)[1]


def next_output(input_path: Path) -> Path:
    versions = []
    for path in DATA.glob("dashboard_kawaii_*.html"):
        match = re.search(r"dashboard_kawaii_(\d+)", path.name)
        if match:
            versions.append(int(match.group(1)))
    return DATA / f"dashboard_kawaii_{max(versions) + 1:03d}_morosidad.html"


def extract_bank() -> list[dict]:
    source = SOURCES / "bcra" / "InfBanc0526.xlsx"
    workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
    sheet = workbook["9"]
    rows = []
    for values in sheet.iter_rows(min_row=9, max_col=9, values_only=True):
        date = iso_month(values[0])
        if date:
            rows.append({
                "date": date,
                "system_private_pct": float(values[1]),
                "companies_pct": float(values[6]),
                "households_pct": float(values[7]),
                "households_personal_cards_pct": float(values[8]),
            })
    if not rows or rows[-1]["date"] != "2026-05-01":
        raise AssertionError("La serie bancaria no termina en mayo de 2026")
    return rows


def extract_pnfc() -> tuple[list[dict], list[dict]]:
    source = SOURCES / "pnfc" / "series-informe-proveedores-no-financieros-credito-junio-2026.xlsx"
    workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
    ratio = workbook["6"]
    dates = [iso_month(cell.value) for cell in ratio[5][2:] if cell.value is not None]
    if not dates or dates[-1] != "2026-02-01":
        raise AssertionError("La serie PNFC no termina en febrero de 2026")

    rows = {date: {"date": date} for date in dates}
    mappings = {
        6: "pnfc_total_pct",
        7: "pnfc_cards_pct",
        8: "pnfc_personal_pct",
        14: "fintech_pct",
        18: "other_card_issuers_pct",
    }
    for row_number, field in mappings.items():
        values = [cell.value for cell in ratio[row_number][2:2 + len(dates)]]
        for date, value in zip(dates, values):
            rows[date][field] = float(value)

    severity = workbook["7"]
    severity_rows = []
    for idx, date in enumerate(dates, start=2):
        irregular = float(severity.cell(6, idx + 1).value)
        follow_up = float(severity.cell(7, idx + 1).value)
        normal = float(severity.cell(8, idx + 1).value)
        severity_rows.append({
            "date": date,
            "universe": "Saldo de cartera PNFC",
            "normal_lt_30d_pct": normal,
            "follow_up_30_90d_pct": follow_up,
            "irregular_gt_90d_pct": irregular,
            "unit": "% del saldo",
        })
    return list(rows.values()), severity_rows


def extract_rates(input_html: str) -> list[dict]:
    match = re.search(
        r'const modern = (\[\{\"date\":\"2019-01-01\".*?\]);',
        input_html,
        flags=re.S,
    )
    if not match:
        raise AssertionError("No se encontró la serie mensual de tasas reales del tab Tasas")
    return json.loads(match.group(1))


def build_data(input_html: str) -> dict:
    bank = extract_bank()
    pnfc, pnfc_severity = extract_pnfc()
    rates = extract_rates(input_html)

    bank_pre = [row for row in bank if row["date"] < "2023-12-01"]
    bank_post = [row for row in bank if row["date"] >= "2023-12-01"]
    bank_mirror = bank_pre[-len(bank_post):]
    bank_hist_mean = statistics.fmean(row["households_pct"] for row in bank_pre)

    cumulative = 0.0
    bank_csv = []
    cumulative_csv = []
    mirror_dates = {row["date"] for row in bank_mirror}
    post_dates = {row["date"] for row in bank_post}
    for row in bank:
        excess = row["households_pct"] - bank_hist_mean
        cumulative += excess
        bank_csv.append({
            **row,
            "historical_mean_pre_shock_pct": bank_hist_mean,
            "excess_vs_historical_mean_pp": excess,
            "cumulative_excess_pp_month": cumulative,
            "mirror_window": "yes" if row["date"] in mirror_dates else "no",
            "post_shock_window": "yes" if row["date"] in post_dates else "no",
            "universe": "Saldo financiado a hogares por entidades financieras",
            "unit": "% del saldo",
        })
        cumulative_csv.append({
            "date": row["date"],
            "households_pct": row["households_pct"],
            "historical_mean_pre_shock_pct": bank_hist_mean,
            "excess_pp": excess,
            "positive_excess_pp": max(0.0, excess),
            "cumulative_excess_pp_month": cumulative,
            "post_shock": "yes" if row["date"] in post_dates else "no",
        })

    product_csv = []
    for row in bank:
        product_csv.extend([
            {
                "date": row["date"], "universe": "Bancos · hogares",
                "category": "Total hogares", "ratio_pct": row["households_pct"],
                "definition": "Cartera irregular / financiaciones a familias",
                "unit_type": "porcentaje de saldo", "interpolated": "no",
            },
            {
                "date": row["date"], "universe": "Bancos · hogares",
                "category": "Personales + tarjetas (combinado)",
                "ratio_pct": row["households_personal_cards_pct"],
                "definition": "Cartera irregular / financiaciones personales y tarjetas combinadas",
                "unit_type": "porcentaje de saldo", "interpolated": "no",
            },
        ])
    for row in pnfc:
        for category, field in [
            ("Total PNFC", "pnfc_total_pct"),
            ("Préstamos personales PNFC", "pnfc_personal_pct"),
            ("Tarjetas PNFC", "pnfc_cards_pct"),
            ("Grupo Fintech", "fintech_pct"),
        ]:
            product_csv.append({
                "date": row["date"], "universe": "PNFC",
                "category": category, "ratio_pct": row[field],
                "definition": "Cartera irregular con mora >90 días / cartera PNFC de la categoría",
                "unit_type": "porcentaje de saldo", "interpolated": "no",
            })

    def window_stats(rows: list[dict], field: str, hist_mean: float) -> dict:
        values = [row[field] for row in rows]
        return {
            "observations": len(rows),
            "start": rows[0]["date"],
            "end": rows[-1]["date"],
            "mora_start_pct": values[0],
            "mora_end_pct": values[-1],
            "average_pct": statistics.fmean(values),
            "maximum_pct": max(values),
            "minimum_pct": min(values),
            "balance_vs_historical_mean_pp_month": sum(v - hist_mean for v in values),
            "months_above_historical_mean": sum(v > hist_mean for v in values),
        }

    bank_before = window_stats(bank_mirror, "households_pct", bank_hist_mean)
    bank_after = window_stats(bank_post, "households_pct", bank_hist_mean)
    bank_diff = (
        bank_after["balance_vs_historical_mean_pp_month"]
        - bank_before["balance_vs_historical_mean_pp_month"]
    )

    pnfc_pre = [row for row in pnfc if row["date"] < "2023-12-01"]
    pnfc_post = [row for row in pnfc if row["date"] >= "2023-12-01"]
    pnfc_mirror = pnfc_pre[-len(pnfc_post):]
    pnfc_hist_mean = statistics.fmean(row["pnfc_total_pct"] for row in pnfc_pre)
    pnfc_before = window_stats(pnfc_mirror, "pnfc_total_pct", pnfc_hist_mean)
    pnfc_after = window_stats(pnfc_post, "pnfc_total_pct", pnfc_hist_mean)
    pnfc_diff = (
        pnfc_after["balance_vs_historical_mean_pp_month"]
        - pnfc_before["balance_vs_historical_mean_pp_month"]
    )

    window_csv = []
    for universe, before, after, diff, mean in [
        ("Bancos · hogares", bank_before, bank_after, bank_diff, bank_hist_mean),
        ("PNFC · total", pnfc_before, pnfc_after, pnfc_diff, pnfc_hist_mean),
    ]:
        for name, values in [("ANTES · espejo", before), ("DESPUÉS · post-shock", after)]:
            window_csv.append({
                "universe": universe, "window": name,
                "historical_mean_pct": mean, **values,
                "differential_post_minus_mirror_pp_month": diff,
                "interpretation": "positive differential = more delinquency / deterioration",
            })

    bank_mirror_aligned = []
    for idx, (before, after) in enumerate(zip(bank_mirror, bank_post), start=1):
        bank_mirror_aligned.append({
            "relative_month": idx,
            "mirror_date": before["date"],
            "mirror_pct": before["households_pct"],
            "mirror_excess_pp": before["households_pct"] - bank_hist_mean,
            "post_date": after["date"],
            "post_pct": after["households_pct"],
            "post_excess_pp": after["households_pct"] - bank_hist_mean,
        })

    rates_by_date = {row["date"]: row.get("bancoReal") for row in rates}
    mora_by_date = {row["date"]: row["households_personal_cards_pct"] for row in bank}
    correlations = []
    for lag in range(7):
        pairs = [
            (rate, mora_by_date[add_month(date, lag)])
            for date, rate in rates_by_date.items()
            if rate is not None and add_month(date, lag) in mora_by_date
        ]
        correlations.append({
            "lag_months": lag,
            "correlation": pearson([x for x, _ in pairs], [y for _, y in pairs]),
            "observations": len(pairs),
            "interpretation": "corr(real personal-loan rate[t-k], bank personal+cards delinquency[t])",
        })
    best_correlation = max(correlations, key=lambda row: abs(row["correlation"]))

    person_snapshots = [
        {"date": "2024-12-01", "provider": "PNFC Fintech", "regular_pct": 86.0,
         "outside_regular_pct": 14.0, "unit_type": "% de personas", "value_type": "derived from 2025 change"},
        {"date": "2024-12-01", "provider": "PNFC tradicionales", "regular_pct": 85.0,
         "outside_regular_pct": 15.0, "unit_type": "% de personas", "value_type": "derived from 2025 change"},
        {"date": "2025-12-01", "provider": "PNFC Fintech", "regular_pct": 79.0,
         "outside_regular_pct": 21.0, "unit_type": "% de personas", "value_type": "official snapshot"},
        {"date": "2025-12-01", "provider": "PNFC tradicionales", "regular_pct": 71.0,
         "outside_regular_pct": 29.0, "unit_type": "% de personas", "value_type": "official snapshot"},
    ]

    mandate_means = []
    for name, start, end, partial in [
        ("Mauricio Macri", "2016-05-01", "2019-11-01", True),
        ("Alberto Fernández", "2019-12-01", "2023-11-01", False),
        ("Javier Milei", "2023-12-01", "2026-05-01", True),
    ]:
        selected = [row["households_pct"] for row in bank if start <= row["date"] <= end]
        mandate_means.append({
            "mandate": name, "start": start, "end": end,
            "average_pct": statistics.fmean(selected), "observations": len(selected),
            "partial": partial,
        })

    latest_bank, nov_bank = bank[-1], next(row for row in bank if row["date"] == "2023-11-01")
    latest_pnfc, nov_pnfc = pnfc[-1], next(row for row in pnfc if row["date"] == "2023-11-01")
    min_bank = min(bank, key=lambda row: row["households_pct"])
    max_bank = max(bank, key=lambda row: row["households_pct"])
    post_positive_area = sum(max(0.0, row["households_pct"] - bank_hist_mean) for row in bank_post)

    kpis = {
        "bank_latest_pct": latest_bank["households_pct"],
        "bank_latest_date": latest_bank["date"],
        "bank_vs_hist_pp": latest_bank["households_pct"] - bank_hist_mean,
        "bank_vs_nov23_pp": latest_bank["households_pct"] - nov_bank["households_pct"],
        "bank_consumer_latest_pct": latest_bank["households_personal_cards_pct"],
        "bank_consumer_vs_nov23_pp": latest_bank["households_personal_cards_pct"] - nov_bank["households_personal_cards_pct"],
        "historical_mean_pct": bank_hist_mean,
        "bank_mirror_pp_month": bank_before["balance_vs_historical_mean_pp_month"],
        "bank_post_pp_month": bank_after["balance_vs_historical_mean_pp_month"],
        "bank_differential_pp_month": bank_diff,
        "bank_post_positive_area_pp_month": post_positive_area,
        "bank_min_pct": min_bank["households_pct"], "bank_min_date": min_bank["date"],
        "bank_max_pct": max_bank["households_pct"], "bank_max_date": max_bank["date"],
        "pnfc_latest_pct": latest_pnfc["pnfc_total_pct"],
        "pnfc_latest_date": latest_pnfc["date"],
        "pnfc_vs_nov23_pp": latest_pnfc["pnfc_total_pct"] - nov_pnfc["pnfc_total_pct"],
        "fintech_latest_pct": latest_pnfc["fintech_pct"],
        "fintech_vs_nov23_pp": latest_pnfc["fintech_pct"] - nov_pnfc["fintech_pct"],
        "pnfc_personal_latest_pct": latest_pnfc["pnfc_personal_pct"],
        "pnfc_personal_vs_nov23_pp": latest_pnfc["pnfc_personal_pct"] - nov_pnfc["pnfc_personal_pct"],
        "pnfc_cards_latest_pct": latest_pnfc["pnfc_cards_pct"],
        "pnfc_cards_vs_nov23_pp": latest_pnfc["pnfc_cards_pct"] - nov_pnfc["pnfc_cards_pct"],
        "pnfc_historical_mean_pct": pnfc_hist_mean,
        "pnfc_mirror_pp_month": pnfc_before["balance_vs_historical_mean_pp_month"],
        "pnfc_post_pp_month": pnfc_after["balance_vs_historical_mean_pp_month"],
        "pnfc_differential_pp_month": pnfc_diff,
        "best_rate_mora_correlation": best_correlation["correlation"],
        "best_rate_mora_lag": best_correlation["lag_months"],
        "fintech_debtors_dec25_million": 6.653465,
        "pnfc_debtors_feb26_million": 6.9,
    }

    return {
        "bank": bank_csv,
        "pnfc": pnfc,
        "pnfc_severity": pnfc_severity,
        "products": product_csv,
        "cumulative": cumulative_csv,
        "windows": window_csv,
        "bank_mirror_aligned": bank_mirror_aligned,
        "correlations": correlations,
        "person_snapshots": person_snapshots,
        "mandate_means": mandate_means,
        "kpis": kpis,
        "meta": {
            "bank_definition": "Cartera irregular / financiaciones a familias; porcentaje del saldo.",
            "pnfc_definition": "Cartera irregular con mora mayor a 90 días / cartera PNFC; porcentaje del saldo.",
            "persons_definition": "Personas deudoras en situación regular o fuera de situación regular; snapshot, porcentaje de personas.",
            "bank_window_months": len(bank_post),
            "pnfc_window_months": len(pnfc_post),
            "interpolation_used": False,
        },
    }


def validate(data: dict) -> dict:
    checks = {}
    rates = [row["households_pct"] for row in data["bank"]]
    rates += [row["ratio_pct"] for row in data["products"]]
    checks["all_rates_between_0_and_100"] = all(0 <= value <= 100 for value in rates)

    bank_rows = [row for row in data["windows"] if row["universe"] == "Bancos · hogares"]
    pnfc_rows = [row for row in data["windows"] if row["universe"] == "PNFC · total"]
    checks["bank_mirror_equal_observations"] = bank_rows[0]["observations"] == bank_rows[1]["observations"]
    checks["pnfc_mirror_equal_observations"] = pnfc_rows[0]["observations"] == pnfc_rows[1]["observations"]
    checks["bank_differential_identity"] = math.isclose(
        bank_rows[1]["balance_vs_historical_mean_pp_month"]
        - bank_rows[0]["balance_vs_historical_mean_pp_month"],
        data["kpis"]["bank_differential_pp_month"], rel_tol=0, abs_tol=1e-10,
    )
    checks["pnfc_differential_identity"] = math.isclose(
        pnfc_rows[1]["balance_vs_historical_mean_pp_month"]
        - pnfc_rows[0]["balance_vs_historical_mean_pp_month"],
        data["kpis"]["pnfc_differential_pp_month"], rel_tol=0, abs_tol=1e-10,
    )
    post = [row for row in data["bank"] if row["post_shock_window"] == "yes"]
    checks["bank_post_balance_is_monthly_sum"] = math.isclose(
        sum(row["excess_vs_historical_mean_pp"] for row in post),
        data["kpis"]["bank_post_pp_month"], rel_tol=0, abs_tol=1e-10,
    )
    checks["latest_kpi_matches_csv"] = math.isclose(
        data["bank"][-1]["households_pct"], data["kpis"]["bank_latest_pct"],
        rel_tol=0, abs_tol=1e-12,
    )
    checks["official_may26_control_12_8"] = abs(data["kpis"]["bank_latest_pct"] - 12.8) < 0.06
    checks["persons_and_balances_have_distinct_units"] = all(
        row["unit_type"] == "% de personas" for row in data["person_snapshots"]
    ) and all(row["unit_type"] == "porcentaje de saldo" for row in data["products"])
    checks["no_interpolation"] = not data["meta"]["interpolation_used"] and all(
        row["interpolated"] == "no" for row in data["products"]
    )
    checks["definitions_separate_bank_pnfc"] = data["meta"]["bank_definition"] != data["meta"]["pnfc_definition"]
    if not all(checks.values()):
        failed = [name for name, ok in checks.items() if not ok]
        raise AssertionError("Fallaron controles de morosidad: " + ", ".join(failed))
    return {"status": "pass", "checks": checks, "generated_at": dt.datetime.now().isoformat(timespec="seconds")}


def update_sources_registry() -> None:
    registry = DATA / "fuentes" / "FUENTES.csv"
    with registry.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        rows = list(reader)
        fields = reader.fieldnames
    assert fields

    def sha(path: Path) -> str:
        digest = hashlib.sha256()
        with path.open("rb") as fh:
            for chunk in iter(lambda: fh.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()

    entries = [
        ("morosidad_bcra_inf_bancos_xlsx", "bcra/InfBanc0526.xlsx", "Informe sobre Bancos · series mayo 2026", "https://www.bcra.gob.ar/archivos/Pdfs/PublicacionesEstadisticas/informes/InfBanc0526.xlsx", "2026-05", "XLSX oficial", "Hoja 9; ratios mensuales de hogares y personales+tarjetas"),
        ("morosidad_bcra_inf_bancos_pdf", "bcra/InfBanc0526.pdf", "Informe sobre Bancos · mayo 2026", "https://www.bcra.gob.ar/archivos/Pdfs/PublicacionesEstadisticas/informes/InfBanc0526.pdf", "2026-05", "PDF oficial", "Definición, control del último dato y período de alivio COVID-19"),
        ("morosidad_bcra_catalogo_series", "bcra/Series_estadisticas.xlsx", "Catálogo de series estadísticas", "https://www.bcra.gob.ar/archivos/Pdfs/PublicacionesEstadisticas/Series_estadisticas.xlsx", "consulta 2026-08-21", "XLSX oficial", "Control de disponibilidad y definiciones; no aporta aperturas bancarias compatibles por producto"),
        ("morosidad_pnfc_series_202606", "pnfc/series-informe-proveedores-no-financieros-credito-junio-2026.xlsx", "Informe PNFC junio 2026 · series", "https://www.bcra.gob.ar/archivos/Pdfs/PublicacionesEstadisticas/informes/series-informe-proveedores-no-financieros-credito-junio-2026.xlsx", "2018-01/2026-02", "XLSX oficial", "Hojas 6 y 7; mora >90 días por producto/proveedor y composición del saldo"),
        ("morosidad_pnfc_pdf_202606", "pnfc/informe-proveedores-no-financieros-credito-junio-2026.pdf", "Informe de Proveedores No Financieros de Crédito · junio 2026", "https://www.bcra.gob.ar/archivos/Pdfs/PublicacionesEstadisticas/informes/informe-proveedores-no-financieros-credito-junio-2026.pdf", "2026-02", "PDF oficial", "Definición, contexto y controles de irregularidad PNFC"),
        ("morosidad_inclusion_xlsx_202604", "central_deudores/anexo-estadistico-informe-inclusion-financiera-2026-04.xlsx", "Anexo estadístico · Informe de Inclusión Financiera abril 2026", "https://www.bcra.gob.ar/archivos/Pdfs/publicacionesestadisticas/informes/anexo-estadistico-informe-inclusion-financiera-2026-04.xlsx", "2025-12", "XLSX oficial", "Cantidad de deudores fintech; gráfico 2.3.5"),
        ("morosidad_inclusion_pdf_202604", "central_deudores/informe-inclusion-financiera-2026-04.pdf", "Informe de Inclusión Financiera · segundo semestre 2025", "https://www.bcra.gob.ar/archivos/Pdfs/publicacionesestadisticas/informes/informe-inclusion-financiera-2026-04.pdf", "2024-12/2025-12", "PDF oficial", "Página 26; porcentaje de personas regulares fintech y PNFC tradicionales"),
    ]
    existing = {row["id"] for row in rows}
    for identifier, relative, title, url, period, file_type, note in entries:
        if identifier in existing:
            continue
        path = SOURCES / relative
        if not path.exists():
            raise FileNotFoundError(path)
        rows.append({
            "id": identifier, "tema": "morosidad", "institucion": "BCRA",
            "titulo": title, "url_original": url,
            "archivo_local": "/data/fuentes/morosidad/" + relative.replace("\\", "/"),
            "fecha_descarga": "2026-08-21", "fecha_publicacion": "",
            "codigo_serie": "", "periodo_utilizado": period, "tipo": file_type,
            "sha256": sha(path), "nota": note,
        })
    with registry.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def build_audit(data: dict, tests: dict) -> str:
    k = data["kpis"]
    return f"""# Auditoría del tab Morosidad

Fecha de corte editorial: 21/08/2026. Último dato bancario: mayo de 2026. Último dato PNFC: febrero de 2026.

## 1. Definiciones y universos

- **Bancos · hogares:** cartera irregular dividida por las financiaciones a familias. Es porcentaje de **saldo**, no de personas. El archivo oficial también publica `Familias · Personales + TC` como agregado; no separa esas dos líneas ni publica allí una serie mensual compatible para hipotecarios y prendarios.
- **PNFC total, personales, tarjetas y grupo Fintech:** cartera con mora **mayor a 90 días** dividida por la cartera de la categoría. Es porcentaje de **saldo**.
- **Personas PNFC:** porcentaje de clientes en situación regular de pago según el Informe de Inclusión Financiera. Se presenta como snapshot anual; no se mezcla con los porcentajes de saldo.
- **Stock no es flujo:** ninguna variación se llama “nuevos morosos”. Puede cambiar por originaciones, pagos, refinanciaciones, castigos, ventas o reclasificaciones.

## 2. Períodos y frecuencia

- Bancos: mensual, mayo de 2016 a mayo de 2026.
- PNFC: mensual, enero de 2018 a febrero de 2026.
- Personas: snapshots de diciembre de 2024 y diciembre de 2025. Los valores 2024 se derivan de los niveles 2025 y de las caídas en regularidad informadas para 2025.
- No se interpoló ninguna observación.

## 3. Promedio histórico y pp-mes

Para cada universo se usa el promedio de todas las observaciones disponibles hasta noviembre de 2023:

`exceso_t = mora_t - promedio_pre_shock`

`saldo_ventana = suma(exceso_t)`

Unidad: **pp-mes**. Positivo significa más morosidad que la norma; negativo, menos.

- Promedio bancario pre-shock: **{fmt(k['historical_mean_pct'])}%**.
- Promedio PNFC pre-shock: **{fmt(k['pnfc_historical_mean_pct'])}%**.

## 4. Ventana espejo

- Bancos: 30 meses espejo (jun-2021 a nov-2023) y 30 meses post-shock (dic-2023 a may-2026).
- PNFC: 27 meses espejo (sep-2021 a nov-2023) y 27 meses post-shock (dic-2023 a feb-2026).
- Diferencial: `saldo_post - saldo_espejo`.
- Diferencial positivo = **empeoró** la morosidad; negativo = mejoró.

Resultados:

- Bancos: antes **{fmt(k['bank_mirror_pp_month'])} pp-mes**, después **{fmt(k['bank_post_pp_month'])} pp-mes**, diferencial **+{fmt(k['bank_differential_pp_month'])} pp-mes**.
- PNFC: antes **{fmt(k['pnfc_mirror_pp_month'])} pp-mes**, después **{fmt(k['pnfc_post_pp_month'])} pp-mes**, diferencial **+{fmt(k['pnfc_differential_pp_month'])} pp-mes**.

## 5. Bancos vs PNFC

Los niveles no se tratan como idénticos. Comparten la idea general de irregularidad sobre saldo, pero PNFC explicita mora >90 días y responde a otro universo de proveedores y clientes. La comparación principal entre ambos se normaliza a nov-2023=100 para observar tendencia.

## 6. Personas y severidad

El informe oficial indica que en diciembre de 2025 estaba regular el 79% de los deudores fintech y el 71% de los deudores PNFC tradicionales. La caída durante 2025 fue de 7 y 14 p.p.; por eso se reconstruyen los snapshots de diciembre de 2024 como 86% y 85%, respectivamente. No se dispone en estas fuentes de una serie histórica mensual comparable de personas en situaciones 1 a 6. La composición mensual PNFC por mora <30, 30–90 y >90 días corresponde a saldos y se conserva separada.

## 7. Cambios regulatorios y cautelas

- El gráfico bancario marca el período de alivio de medidas financieras por COVID-19 que el propio BCRA sombrea en el Informe sobre Bancos.
- Desde julio de 2024, la Central de Deudores elevó el umbral mínimo informado de $1.000 a $25.000. Ese quiebre afecta conteos de personas; por eso no se construyó una serie mensual espuria a partir de esos conteos.
- Hipotecarios y prendarios no se incorporan como series porque el workbook mensual vigente no expone aperturas compatibles y continuas. El informe narrativo los menciona como impulsores en algunos meses, pero eso no alcanza para fabricar una serie.

## 8. Tasas reales y correlación

Se explora `corr(tasa_real_personales[t-k], mora_personales_más_tarjetas[t])` para k=0…6. La mayor correlación absoluta es **r={k['best_rate_mora_correlation']:.3f}** con **{k['best_rate_mora_lag']} meses** de rezago. Es sincronía temporal, no prueba causal.

## 9. Resultado principal

La mora bancaria de hogares alcanzó **{fmt(k['bank_latest_pct'], 1)}%** en mayo de 2026, **+{fmt(k['bank_vs_hist_pp'])} p.p.** sobre su promedio pre-shock y **+{fmt(k['bank_vs_nov23_pp'])} p.p.** frente a noviembre de 2023. La ventana post-shock resultó **{fmt(k['bank_differential_pp_month'])} pp-mes más desfavorable** que su espejo. PNFC también empeoró frente a noviembre de 2023, aunque su saldo acumulado en ambas ventanas sigue debajo de un promedio histórico elevado por años de mora muy alta.

## 10. Controles automáticos

Estado: **{tests['status'].upper()}**. Controles aprobados: {len(tests['checks'])}/{len(tests['checks'])}.
"""


CSS = r"""
/* MOROSIDAD_TAB_VERSION:1 */
.mor-shell{display:grid;gap:16px;color:#553962;min-width:0}.mor-shell>*,.mor-panel,.mor-grid-2>div{min-width:0}.mor-hero,.mor-panel{border:1px solid #e2cdec;border-radius:24px;background:rgba(255,255,255,.9);box-shadow:0 10px 28px rgba(95,61,111,.07);padding:20px;box-sizing:border-box}.mor-hero{background:linear-gradient(135deg,rgba(255,249,252,.97),rgba(247,255,251,.95) 55%,rgba(250,247,255,.96))}.mor-head{display:flex;justify-content:space-between;gap:18px;align-items:flex-start}.mor-head h2,.mor-head h3{margin:0;color:#52306f}.mor-head h2{font-size:28px}.mor-sub,.mor-note{margin:7px 0 0;color:#735d7a;line-height:1.55}.mor-badge{display:inline-block;margin-bottom:7px;padding:5px 10px;border-radius:999px;background:#f6edf9;border:1px solid #ddcce8;font-size:10px;font-weight:900;text-transform:uppercase;color:#775383}.mor-kpis{display:grid;grid-template-columns:repeat(7,minmax(0,1fr));gap:10px;margin-top:17px}.mor-kpi{min-width:0;padding:13px;border:1px solid #e8dcea;border-radius:17px;background:#fff}.mor-kpi small{display:block;font-size:8.5px;font-weight:950;text-transform:uppercase;color:#896f8e}.mor-kpi strong{display:block;margin:6px 0 4px;font-size:23px;line-height:1;color:#ae426d}.mor-kpi span{display:block;font-size:9px;line-height:1.35;color:#77657d}.mor-kpi.good strong{color:#2e8a6b}.mor-kpi.context strong{color:#5c48a5}.mor-reading{margin-top:14px;padding:13px 15px;border-left:4px solid #a668c1;border-radius:13px;background:#fbf8ff;line-height:1.58;font-size:11px}.mor-reading b{color:#563774}.mor-chart{width:100%;max-width:100%;height:410px;min-height:340px;min-width:0}.mor-chart .plot-container,.mor-chart .svg-container{max-width:100%!important}.mor-chart.tall{height:480px}.mor-grid-2{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:16px}.mor-grid-3{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:11px}.mor-stat{padding:14px;border:1px solid #e4d7e7;border-radius:16px;background:#fff}.mor-stat small{display:block;font-size:8px;font-weight:950;text-transform:uppercase;color:#89738e}.mor-stat strong{display:block;margin:6px 0;font-size:24px;color:#a9426d}.mor-stat.good strong{color:#30876d}.mor-stat.context strong{color:#5c47a4}.mor-stat span{font-size:9px;line-height:1.35;color:#746279}.mor-universe{margin:10px 0;padding:9px 12px;border:1px dashed #d8c6e2;border-radius:13px;background:#fcf9ff;font-size:10px;line-height:1.45;color:#735c7d}.mor-plain{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:10px;margin-top:12px}.mor-plain>div{padding:12px;border:1px solid #e7dce9;border-radius:14px;background:#fff;font-size:10px;line-height:1.5}.mor-plain b{display:block;margin-bottom:4px;color:#563974}.mor-downloads,.mor-links{display:flex;gap:8px;flex-wrap:wrap;margin-top:12px}.mor-btn{border:1px solid #d9c3e4;border-radius:999px;background:#fff;padding:9px 13px;color:#684379;font:inherit;font-size:10px;font-weight:900;cursor:pointer}.mor-btn:hover{background:#f8eefb}.mor-table-wrap{overflow-x:auto;margin-top:12px;min-width:0;max-width:100%}.mor-table-wrap table{width:100%;border-collapse:separate;border-spacing:0;font-size:10px}.mor-table-wrap th,.mor-table-wrap td{padding:9px;border-bottom:1px solid #eee3f0;text-align:left;white-space:nowrap}.mor-table-wrap th{background:#f8f2fa;color:#775481;text-transform:uppercase;font-size:8px}.mor-caution{margin-top:12px;padding:12px;border:1px solid #f1d09b;border-radius:14px;background:#fffaf1;color:#765b35;font-size:10px;line-height:1.5}.mor-methods{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:10px}.mor-methods>div{padding:13px;border-radius:15px;border:1px solid #e6d9e9;background:#fff;font-size:10px;line-height:1.5}.mor-methods b{display:block;color:#5a3a6e;margin-bottom:5px}.mor-milei-info{margin:14px 0;padding:14px 16px;border:1px solid #e5cadd;border-radius:18px;background:#fff8fb}.mor-milei-info h3{margin:0 0 7px;color:#653b62}.mor-milei-info p{margin:0;color:#765f73;line-height:1.5;font-size:10px}.mor-inline-kpis{display:flex;gap:9px;flex-wrap:wrap;margin:10px 0}.mor-inline-kpis span{padding:7px 10px;border-radius:12px;background:#fff;border:1px solid #eadce7;font-size:9px;font-weight:850}.mor-rates-jump{margin:0 0 14px;padding:10px 14px;border:1px solid #d9c4e6;border-radius:15px;background:#fcf8ff;display:flex;align-items:center;justify-content:space-between;gap:12px;font-size:10px;color:#705578}.mor-rates-jump b{color:#54336c}.mor-rates-jump button{white-space:nowrap}
@media(max-width:1250px){.mor-kpis{grid-template-columns:repeat(4,minmax(0,1fr))}}
@media(max-width:900px){.mor-grid-2,.mor-methods{grid-template-columns:1fr}.mor-kpis{grid-template-columns:repeat(2,minmax(0,1fr))}.mor-head{display:block}.mor-chart,.mor-chart.tall{height:420px}.mor-plain{grid-template-columns:1fr}}
@media(max-width:520px){.mor-hero,.mor-panel{padding:14px;border-radius:18px}.mor-head h2{font-size:22px}.mor-kpis{grid-template-columns:1fr 1fr;gap:8px}.mor-kpi{padding:10px}.mor-kpi strong{font-size:19px}.mor-chart,.mor-chart.tall{height:390px;min-height:330px}.mor-grid-3{grid-template-columns:1fr}.mor-rates-jump{align-items:flex-start;flex-direction:column}.mor-table-wrap{margin-left:-4px;margin-right:-4px}}
"""


SECTION = r"""
  <!-- MOROSIDAD_TAB_VERSION:1 -->
  <section id="tab-morosidad" class="tab-panel">
    <div class="mor-shell">
      <section class="mor-hero">
        <div class="mor-head"><div><span class="mor-badge">BCRA · saldos mensuales · personas sólo en snapshots oficiales</span><h2>Morosidad · ¿la gente puede pagar sus deudas? ♡</h2><p class="mor-sub">Mide la parte del crédito que está en situación irregular. Separamos bancos, PNFC/Fintech y personas porque no usan exactamente el mismo universo ni denominador.</p></div></div>
        <div class="mor-kpis">
          <div class="mor-kpi"><small>Morosidad hogares</small><strong>12,8%</strong><span>mayo de 2026 · % del saldo bancario</span></div>
          <div class="mor-kpi"><small>Vs promedio histórico</small><strong>+9,42 pp</strong><span>más mora que la norma pre-shock</span></div>
          <div class="mor-kpi"><small>Personales + tarjetas</small><strong>14,5%</strong><span>bancos · ambas líneas combinadas</span></div>
          <div class="mor-kpi context"><small>Ventana espejo</small><strong>+5,33</strong><span>pp-mes contra la norma</span></div>
          <div class="mor-kpi"><small>Post-shock</small><strong>+59,04</strong><span>pp-mes contra la norma</span></div>
          <div class="mor-kpi"><small>Diferencial</small><strong>+53,71</strong><span>pp-mes · empeoró vs espejo</span></div>
          <div class="mor-kpi"><small>Fintech</small><strong>26,2%</strong><span>feb-2026 · mora &gt;90d sobre saldo PNFC Fintech</span></div>
        </div>
        <div class="mor-reading"><b>Respuesta corta:</b> la capacidad de pago medida por la mora bancaria empeoró frente a noviembre de 2023, frente a su promedio histórico y frente a una ventana anterior de igual duración. En PNFC también empeoró la comparación entre ventanas, aunque ambas acumulan menos mora que el promedio pre-shock —un promedio elevado por años anteriores muy malos—. Esto describe evolución; no demuestra qué la causó.</div>
      </section>

      <section class="mor-panel"><div class="mor-head"><div><h3>1 · Morosidad bancaria de hogares</h3><p class="mor-note">De cada $100 financiados a hogares, $12,80 estaban en situación irregular en mayo de 2026. La línea combinada de personales + tarjetas llegó a $14,52 por cada $100 de esas dos carteras.</p></div></div><div id="morBankChart" class="mor-chart tall"></div><div id="morMandateTable" class="mor-table-wrap"></div><div class="mor-caution"><b>Zona gris COVID-19:</b> el propio BCRA identifica el período de alivio de medidas financieras. Puede afectar la clasificación observada; no se interpreta como una mejora “natural” de pago.</div></section>

      <section class="mor-panel"><div class="mor-head"><div><h3>2 · Por tipo de crédito: aperturas realmente publicadas</h3><p class="mor-note">El workbook bancario vigente combina personales y tarjetas; PNFC sí las separa. No inventamos series de hipotecarios o prendarios a partir de menciones narrativas.</p></div></div><div class="mor-grid-2"><div><div class="mor-universe"><b>Bancos · porcentaje del saldo.</b> Hogares totales y personales + tarjetas combinados.</div><div id="morBankProductsChart" class="mor-chart"></div></div><div><div class="mor-universe"><b>PNFC · mora &gt;90 días · porcentaje del saldo.</b> Personales y tarjetas separados.</div><div id="morPnfcProductsChart" class="mor-chart"></div></div></div><div class="mor-plain"><div><b>Mayor tasa actual</b>Personales PNFC: 34,1% en feb-2026.</div><div><b>Mayor aumento desde nov-2023</b>Tarjetas PNFC: +15,2 p.p.; personales: +12,3 p.p.</div><div><b>Bancos · línea disponible</b>Personales + tarjetas combinadas: +12,13 p.p. desde nov-2023.</div></div></section>

      <section class="mor-panel"><div class="mor-head"><div><h3>3 · Bancos vs PNFC / Fintech</h3><p class="mor-note">Normalizamos cada serie a nov-2023=100 para comparar dirección y velocidad, no nivel absoluto.</p></div></div><div class="mor-universe"><b>Universos distintos · comparar tendencia, no nivel absoluto.</b> Bancos usa cartera irregular de hogares; PNFC define irregularidad como mora &gt;90 días. Fintech es un grupo dentro de PNFC.</div><div id="morCompareChart" class="mor-chart"></div><div class="mor-grid-3"><div class="mor-stat"><small>PNFC total</small><strong>26,9%</strong><span>+16,1 p.p. vs nov-2023</span></div><div class="mor-stat"><small>Fintech</small><strong>26,2%</strong><span>+4,2 p.p. vs nov-2023</span></div><div class="mor-stat context"><small>Clientes PNFC</small><strong>6,9 M</strong><span>personas · feb-2026</span></div></div></section>

      <section class="mor-panel"><div class="mor-head"><div><h3>4 · Saldo acumulado contra la norma histórica</h3><p class="mor-note"><code>exceso_t = mora_t − promedio pre-shock</code>. Sumarlo produce pp-mes: intensidad × duración del desvío. Positivo es peor para el deudor.</p></div></div><div id="morCumulativeChart" class="mor-chart"></div><div class="mor-grid-3"><div class="mor-stat context"><small>Norma pre-shock</small><strong>3,38%</strong><span>promedio may-2016→nov-2023</span></div><div class="mor-stat"><small>Área positiva post</small><strong>69,78</strong><span>pp-mes de mora extraordinaria</span></div><div class="mor-stat"><small>Último vs norma</small><strong>+9,42 pp</strong><span>mayo de 2026</span></div></div></section>

      <section class="mor-panel"><div class="mor-head"><div><h3>5 · Ventana espejo: ANTES, DESPUÉS y DIFERENCIAL</h3><p class="mor-note">30 meses por ventana en bancos. El diferencial es post menos espejo: positivo significa más morosidad y, por lo tanto, empeoramiento.</p></div></div><div id="morMirrorChart" class="mor-chart"></div><div class="mor-grid-3"><div class="mor-stat context"><small>ANTES · espejo</small><strong>+5,33</strong><span>pp-mes · jun-2021→nov-2023</span></div><div class="mor-stat"><small>DESPUÉS · post-shock</small><strong>+59,04</strong><span>pp-mes · dic-2023→may-2026</span></div><div class="mor-stat"><small>DIFERENCIAL</small><strong>+53,71</strong><span>pp-mes · empeoró</span></div></div><div class="mor-universe"><b>PNFC, su propia ventana de 27 meses:</b> espejo −153,68 pp-mes; post −147,08; diferencial <b>+6,60 pp-mes = empeoró</b>. Ambas ventanas quedan debajo de su promedio histórico; eso no contradice que la segunda sea menos favorable.</div></section>

      <section class="mor-panel"><div class="mor-head"><div><h3>6 · Personas: snapshots oficiales, no pesos</h3><p class="mor-note">El Informe de Inclusión Financiera permite una foto de personas regulares y fuera de situación regular. No permite reconstruir una serie mensual comparable de situaciones 1 a 6.</p></div></div><div id="morPersonsChart" class="mor-chart"></div><div class="mor-grid-3"><div class="mor-stat context"><small>Fintech · dic-2025</small><strong>21 de 100</strong><span>deudores fuera de situación regular</span></div><div class="mor-stat"><small>PNFC tradicional</small><strong>29 de 100</strong><span>fuera de situación regular</span></div><div class="mor-stat context"><small>Crédito digital</small><strong>6,7 M</strong><span>personas con PNFC Fintech · dic-2025</span></div></div><div class="mor-caution">El umbral mínimo reportado en la Central de Deudores cambió en julio de 2024, de $1.000 a $25.000. Por eso usamos snapshots oficiales y no llamamos “nuevos morosos” a variaciones de conteos.</div></section>

      <section class="mor-panel"><div class="mor-head"><div><h3>7 · Tasas reales vs mora: exploración con rezagos</h3><p class="mor-note">Correlación entre la tasa real mensual de préstamos personales y la mora bancaria combinada de personales + tarjetas. k indica cuántos meses se adelanta la tasa respecto de la mora.</p></div></div><div id="morCorrelationChart" class="mor-chart"></div><div class="mor-grid-2"><div class="mor-stat context"><small>Correlación máxima</small><strong>r = 0,488</strong><span>con tasa adelantada 6 meses · 83 observaciones</span></div><div class="mor-caution"><b>Correlación temporal ≠ causalidad.</b> El resultado no prueba que la tasa haya causado la mora; salarios, actividad, originación, refinanciaciones y composición también pueden intervenir.</div></div></section>

      <section class="mor-panel"><div class="mor-head"><div><h3>Lectura automática · en criollo</h3></div></div><div class="mor-reading"><b>¿Está peor que antes?</b> Sí en el indicador bancario: la mora de hogares pasó de 2,70% en nov-2023 a 12,80% en may-2026 (+10,10 p.p.). El post-shock acumuló 53,71 pp-mes más mora contra la norma que la ventana espejo. <b>¿Dónde se concentró?</b> La línea bancaria de personales + tarjetas llegó a 14,52%; en PNFC, personales tiene el nivel más alto (34,1%), mientras tarjetas fue la que más aumentó desde nov-2023 (+15,2 p.p.). <b>¿Y Fintech?</b> Su ratio de saldo irregular subió 4,2 p.p.; entre personas, la proporción fuera de situación regular pasó de 14% a 21% durante 2025. El deterioro aparece tanto en saldos como en snapshots de personas, pero las definiciones no permiten sumarlos.</div>
        <div class="mor-methods"><div><b>Saldo vs cambio</b>Un saldo positivo en pp-mes es desfavorable: hubo más mora que la norma. Un diferencial positivo también es desfavorable: empeoró frente al espejo.</div><div><b>Lo que no contamos</b>Más deuda no implica automáticamente más mora. Variar el stock irregular tampoco equivale a contar nuevas personas morosas.</div><div><b>Lo que falta</b>No hay serie mensual compatible de hipotecarios/prendarios ni de personas en situaciones 1–6 en los anexos usados. Queda documentado, no rellenado.</div></div>
        <div class="mor-downloads"><button class="mor-btn" onclick="downloadMorCsv('households')">CSV · hogares</button><button class="mor-btn" onclick="downloadMorCsv('products')">CSV · productos</button><button class="mor-btn" onclick="downloadMorCsv('pnfc')">CSV · PNFC</button><button class="mor-btn" onclick="downloadMorCsv('mirror')">CSV · ventana espejo</button><button class="mor-btn" onclick="downloadMorCsv('persons')">CSV · personas</button></div>
        <div class="mor-links"><button class="mor-btn" onclick="activateTabAndScroll('tab-rates')">Tasas e inflación</button><button class="mor-btn" onclick="activateTabAndScroll('tab-power')">Poder adquisitivo</button><button class="mor-btn" onclick="activateTabAndScroll('tab-work')">Trabajo</button><button class="mor-btn" onclick="activateTabAndScroll('tab-consumption')">Consumo</button><button class="mor-btn" onclick="activateTabAndScroll('tab-emae')">Actividad real / EMAE</button></div>
      </section>
    </div>
  </section>
"""


RATES_JUMP = r"""
    <div class="mor-rates-jump"><span><b>¿El crédito caro terminó en mora?</b> Mirá la evolución bancaria, PNFC/Fintech y la correlación exploratoria con rezagos.</span><button class="mor-btn" type="button" onclick="activateTabAndScroll('tab-morosidad')">Ver Morosidad →</button></div>
"""


MILEI_INFO = r"""
    <section class="mor-milei-info" id="mileiMorosityInfo"><h3>Estrés financiero · morosidad</h3><p>Indicador de capacidad de pago; <b>no se suma al total monetario</b> porque convertir cartera irregular en pérdida duplicaría o exageraría costos.</p><div class="mor-inline-kpis"><span>Hogares: 12,8%</span><span>vs nov-2023: +10,10 pp</span><span>post vs espejo: +53,71 pp-mes</span><span>Fintech: 26,2%</span></div><button class="mor-btn" onclick="activateTabAndScroll('tab-morosidad')">Ver Morosidad →</button></section>
"""


JS = r"""
<script>
const MOR_DATA=__MOR_DATA__;
const MOR_CSVS=__MOR_CSVS__;
let morRendered=false;
function morFmt(v,d=2){return Number(v).toLocaleString('es-AR',{minimumFractionDigits:d,maximumFractionDigits:d})}
function morLayout(title,ytitle,extra={}){const mobile=window.innerWidth<600;return {title:{text:title,font:{size:mobile?13:16,color:'#54366b'}},paper_bgcolor:'rgba(0,0,0,0)',plot_bgcolor:'rgba(255,255,255,.62)',font:{family:'inherit',color:'#624b6b',size:mobile?9:11},margin:{l:mobile?48:62,r:mobile?28:35,t:mobile?82:62,b:52},hovermode:'x unified',legend:{orientation:mobile?'v':'h',y:mobile?1.2:1.08,x:0,font:{size:mobile?9:11}},xaxis:{gridcolor:'#eee5f2',tickformat:'%Y',automargin:true},yaxis:{title:ytitle,gridcolor:'#eadff0',zerolinecolor:'#bca9c7',automargin:true},...extra}}
function morBands(){return {shapes:[{type:'rect',xref:'x',yref:'paper',x0:'2016-05-01',x1:'2019-12-01',y0:0,y1:1,fillcolor:'#fff1da',opacity:.32,line:{width:0},layer:'below'},{type:'rect',xref:'x',yref:'paper',x0:'2019-12-01',x1:'2023-12-01',y0:0,y1:1,fillcolor:'#e5f2fb',opacity:.34,line:{width:0},layer:'below'},{type:'rect',xref:'x',yref:'paper',x0:'2023-12-01',x1:'2026-06-01',y0:0,y1:1,fillcolor:'#fde8f0',opacity:.38,line:{width:0},layer:'below'},{type:'rect',xref:'x',yref:'paper',x0:'2020-03-01',x1:'2021-04-01',y0:0,y1:1,fillcolor:'#9b9b9b',opacity:.14,line:{width:0},layer:'below'},{type:'line',xref:'x',yref:'paper',x0:'2023-12-10',x1:'2023-12-10',y0:0,y1:1,line:{color:'#d84f85',width:2,dash:'dot'}}],annotations:[{xref:'x',yref:'paper',x:'2018-02-01',y:1.02,text:'Macri · tramo parcial',showarrow:false,font:{size:9,color:'#7c617f'}},{xref:'x',yref:'paper',x:'2021-12-01',y:1.02,text:'Alberto Fernández',showarrow:false,font:{size:9,color:'#7c617f'}},{xref:'x',yref:'paper',x:'2025-03-01',y:1.02,text:'Milei · parcial',showarrow:false,font:{size:9,color:'#7c617f'}},{xref:'x',yref:'paper',x:'2020-09-01',y:.12,text:'alivio COVID-19',showarrow:false,font:{size:8,color:'#777'}}]}}
function renderMor(){if(!window.Plotly)return;const b=MOR_DATA.bank,p=MOR_DATA.pnfc,k=MOR_DATA.kpis;const bx=b.map(r=>r.date),px=p.map(r=>r.date);const deco=morBands();
 Plotly.react('morBankChart',[{x:bx,y:b.map(r=>r.households_pct),name:'Hogares',line:{color:'#5d4bc4',width:3}},{x:bx,y:b.map(r=>r.households_personal_cards_pct),name:'Personales + tarjetas',line:{color:'#e65e92',width:2.5}},{x:bx,y:b.map(()=>k.historical_mean_pct),name:'Promedio pre-shock',line:{color:'#2f8b70',width:2,dash:'dash'}}],morLayout('Cartera irregular / financiaciones a hogares','% del saldo',{...deco,annotations:[...deco.annotations,{x:k.bank_latest_date,y:k.bank_latest_pct,text:'Último '+morFmt(k.bank_latest_pct,1)+'%',showarrow:true,arrowcolor:'#d84f85',bgcolor:'#fff7fb',bordercolor:'#e7b8ca'}]}),{responsive:true,displaylogo:false});
 Plotly.react('morBankProductsChart',[{x:bx,y:b.map(r=>r.households_pct),name:'Hogares total',line:{color:'#6047bf',width:3}},{x:bx,y:b.map(r=>r.households_personal_cards_pct),name:'Personales + tarjetas',line:{color:'#e55f91',width:3}}],morLayout('Bancos · aperturas compatibles','% del saldo',deco),{responsive:true,displaylogo:false});
 Plotly.react('morPnfcProductsChart',[{x:px,y:p.map(r=>r.pnfc_total_pct),name:'PNFC total',line:{color:'#6047bf',width:2}},{x:px,y:p.map(r=>r.pnfc_personal_pct),name:'Personales',line:{color:'#e55f91',width:3}},{x:px,y:p.map(r=>r.pnfc_cards_pct),name:'Tarjetas',line:{color:'#ef8b45',width:3}}],morLayout('PNFC · por tipo de asistencia','% del saldo con mora >90d',{shapes:[{type:'line',xref:'x',yref:'paper',x0:'2023-12-10',x1:'2023-12-10',y0:0,y1:1,line:{color:'#d84f85',dash:'dot'}}]}),{responsive:true,displaylogo:false});
 const postB=b.filter(r=>r.date>='2023-11-01'),postP=p.filter(r=>r.date>='2023-11-01');const baseB=postB[0].households_pct,baseP=postP[0].pnfc_total_pct,baseF=postP[0].fintech_pct;
 Plotly.react('morCompareChart',[{x:postB.map(r=>r.date),y:postB.map(r=>r.households_pct/baseB*100),name:'Bancos · hogares',line:{color:'#6047bf',width:3}},{x:postP.map(r=>r.date),y:postP.map(r=>r.pnfc_total_pct/baseP*100),name:'PNFC total',line:{color:'#ef8b45',width:3}},{x:postP.map(r=>r.date),y:postP.map(r=>r.fintech_pct/baseF*100),name:'Fintech',line:{color:'#df588c',width:3,dash:'dot'}}],morLayout('Tendencia normalizada','nov-2023 = 100',{shapes:[{type:'line',xref:'paper',x0:0,x1:1,y0:100,y1:100,line:{color:'#318a6e',dash:'dash'}}]}),{responsive:true,displaylogo:false});
 const c=MOR_DATA.cumulative;Plotly.react('morCumulativeChart',[{x:c.map(r=>r.date),y:c.map(r=>r.excess_pp),name:'Desvío mensual',type:'bar',marker:{color:c.map(r=>r.excess_pp>0?'rgba(216,76,130,.48)':'rgba(49,139,111,.45)')}},{x:c.map(r=>r.date),y:c.map(r=>r.cumulative_excess_pp_month),name:'Saldo acumulado',yaxis:'y2',line:{color:'#5b3fa5',width:3}}],morLayout('Desvío mensual y saldo acumulado','desvío mensual · pp',{yaxis2:{title:'saldo acumulado · pp-mes',overlaying:'y',side:'right',gridcolor:'rgba(0,0,0,0)',automargin:true},shapes:[{type:'line',xref:'x',yref:'paper',x0:'2023-12-10',x1:'2023-12-10',y0:0,y1:1,line:{color:'#d84f85',dash:'dot'}}],barmode:'relative'}),{responsive:true,displaylogo:false});
 const m=MOR_DATA.bank_mirror_aligned;Plotly.react('morMirrorChart',[{x:m.map(r=>r.relative_month),y:m.map(r=>r.mirror_excess_pp),name:'ANTES · espejo',line:{color:'#3b8c70',width:3}},{x:m.map(r=>r.relative_month),y:m.map(r=>r.post_excess_pp),name:'DESPUÉS · post-shock',line:{color:'#d84f85',width:3}}],morLayout('Desvío de mora contra la norma en cada mes relativo','puntos porcentuales',{xaxis:{title:'mes relativo de cada ventana',gridcolor:'#eee5f2'},shapes:[{type:'line',xref:'paper',x0:0,x1:1,y0:0,y1:0,line:{color:'#8f7b99',dash:'dash'}}]}),{responsive:true,displaylogo:false});
 const s=MOR_DATA.person_snapshots,labels=s.map(r=>(r.provider.includes('Fintech')?'Fintech':'PNFC tradicional')+' · '+r.date.slice(0,4));Plotly.react('morPersonsChart',[{x:labels,y:s.map(r=>r.regular_pct),name:'Situación regular',type:'bar',marker:{color:'#63b493'}},{x:labels,y:s.map(r=>r.outside_regular_pct),name:'Fuera de situación regular',type:'bar',marker:{color:'#df5d8d'}}],morLayout('Composición de personas deudoras · snapshots','% de personas',{barmode:'stack',xaxis:{gridcolor:'#eee5f2',tickangle:window.innerWidth<600?-18:0},yaxis:{title:'% de personas',range:[0,100],gridcolor:'#eadff0'}}),{responsive:true,displaylogo:false});
 const corr=MOR_DATA.correlations;Plotly.react('morCorrelationChart',[{x:corr.map(r=>r.lag_months),y:corr.map(r=>r.correlation),type:'bar',marker:{color:corr.map(r=>r.lag_months===k.best_rate_mora_lag?'#d75587':'#8a73c7')},text:corr.map(r=>morFmt(r.correlation,3)),textposition:'outside',cliponaxis:false}],morLayout('Correlación por rezago de la tasa real','coeficiente r',{xaxis:{title:'rezago k · meses',dtick:1,gridcolor:'#eee5f2'},yaxis:{title:'coeficiente r',range:[0,.62],gridcolor:'#eadff0'}}),{responsive:true,displaylogo:false});
 document.getElementById('morMandateTable').innerHTML='<table><thead><tr><th>Mandato</th><th>Período observado</th><th>Promedio de mora</th><th>Observaciones</th><th>Alcance</th></tr></thead><tbody>'+MOR_DATA.mandate_means.map(r=>`<tr><td><b>${r.mandate}</b></td><td>${r.start.slice(0,7)} → ${r.end.slice(0,7)}</td><td>${morFmt(r.average_pct,2)}%</td><td>${r.observations}</td><td>${r.partial?'parcial según disponibilidad':'completo'}</td></tr>`).join('')+'</tbody></table>';morRendered=true;}
function downloadMorCsv(key){const item=MOR_CSVS[key];if(!item)return;const blob=new Blob(['\ufeff'+item.content],{type:'text/csv;charset=utf-8'});const a=document.createElement('a');a.href=URL.createObjectURL(blob);a.download=item.filename;document.body.appendChild(a);a.click();a.remove();setTimeout(()=>URL.revokeObjectURL(a.href),1000)}
document.querySelector('[data-tab="tab-morosidad"]')?.addEventListener('click',()=>requestAnimationFrame(renderMor));
window.addEventListener('resize',()=>{if(!morRendered)return;['morBankChart','morBankProductsChart','morPnfcProductsChart','morCompareChart','morCumulativeChart','morMirrorChart','morPersonsChart','morCorrelationChart'].forEach(id=>{const el=document.getElementById(id);if(el&&window.Plotly)Plotly.Plots.resize(el)})});
</script>
"""


def integrate_html(input_html: str, data: dict, csvs: dict[str, dict]) -> str:
    if "MOROSIDAD_TAB_VERSION" in input_html:
        raise AssertionError("El dashboard de entrada ya contiene el tab Morosidad")
    html = input_html
    html = html.replace(
        '<button class="tab-btn" data-tab="tab-emae">Actividad real · ¿crecimiento o rebote?</button>',
        '<button class="tab-btn" data-tab="tab-emae">Actividad real · ¿crecimiento o rebote?</button>\n    <button class="tab-btn" data-tab="tab-morosidad">Morosidad · ¿la gente puede pagar sus deudas?</button>',
        1,
    )
    if 'data-tab="tab-morosidad"' not in html:
        raise AssertionError("No se pudo insertar el botón de navegación")
    html = html.replace("</style>", CSS + "\n</style>", 1)
    html = html.replace(
        '  <section id="tab-fiscal" class="tab-panel">',
        SECTION + '\n\n  <section id="tab-fiscal" class="tab-panel">',
        1,
    )
    rates_start = html.find('  <section id="tab-rates" class="tab-panel">')
    if rates_start < 0:
        raise AssertionError("No se encontró el tab Tasas")
    rates_grid = html.find('    <div class="grid">', rates_start)
    if rates_grid < 0:
        raise AssertionError("No se encontró la grilla inicial del tab Tasas")
    html = html[:rates_grid] + RATES_JUMP + "\n" + html[rates_grid:]
    html = html.replace(
        '    <div id="mileiCostHero" class="milei-cost-hero"></div>',
        '    <div id="mileiCostHero" class="milei-cost-hero"></div>\n' + MILEI_INFO,
        1,
    )
    js = JS.replace("__MOR_DATA__", json.dumps(data, ensure_ascii=False, separators=(",", ":")))
    js = js.replace("__MOR_CSVS__", json.dumps(csvs, ensure_ascii=False, separators=(",", ":")))
    html = html.replace("</body>", js + "\n</body>", 1)
    return html


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path, default=None)
    parser.add_argument("--output", type=Path, default=None)
    args = parser.parse_args()
    input_path = args.input.resolve() if args.input else latest_dashboard()
    output_path = args.output.resolve() if args.output else next_output(input_path)
    if input_path == output_path:
        raise SystemExit("El archivo de salida no puede sobrescribir la entrada")
    input_html = input_path.read_text(encoding="utf-8")
    data = build_data(input_html)
    tests = validate(data)

    write_csv(DERIVED / "morosidad_hogares.csv", data["bank"])
    write_csv(DERIVED / "morosidad_por_producto.csv", data["products"])
    write_csv(DERIVED / "morosidad_pnfc.csv", data["pnfc"])
    write_csv(DERIVED / "morosidad_ventana_espejo.csv", data["windows"])
    write_csv(DERIVED / "morosidad_saldo_acumulado.csv", data["cumulative"])
    write_csv(DERIVED / "situacion_deudores_personas.csv", data["person_snapshots"])
    write_csv(DERIVED / "morosidad_pnfc_severidad_saldo.csv", data["pnfc_severity"])
    write_csv(DERIVED / "correlacion_tasa_real_mora.csv", data["correlations"])
    (DERIVED / "TESTS_MOROSIDAD.json").write_text(
        json.dumps(tests, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (DERIVED / "AUDITORIA_MOROSIDAD.md").write_text(
        build_audit(data, tests), encoding="utf-8"
    )
    update_sources_registry()

    csvs = {
        "households": {"filename": "morosidad_hogares.csv", "content": read_csv_text(DERIVED / "morosidad_hogares.csv")},
        "products": {"filename": "morosidad_por_producto.csv", "content": read_csv_text(DERIVED / "morosidad_por_producto.csv")},
        "pnfc": {"filename": "morosidad_pnfc.csv", "content": read_csv_text(DERIVED / "morosidad_pnfc.csv")},
        "mirror": {"filename": "morosidad_ventana_espejo.csv", "content": read_csv_text(DERIVED / "morosidad_ventana_espejo.csv")},
        "persons": {"filename": "situacion_deudores_personas.csv", "content": read_csv_text(DERIVED / "situacion_deudores_personas.csv")},
    }
    output_html = integrate_html(input_html, data, csvs)
    output_path.write_text(output_html, encoding="utf-8")
    print(json.dumps({
        "input": str(input_path), "output": str(output_path),
        "bank_latest_pct": data["kpis"]["bank_latest_pct"],
        "bank_differential_pp_month": data["kpis"]["bank_differential_pp_month"],
        "pnfc_differential_pp_month": data["kpis"]["pnfc_differential_pp_month"],
        "tests": tests["status"],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
